"""Garde-fous transversaux de l'administration RH."""

import io

from django.contrib import admin
from django.contrib import messages
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import load_workbook

from .models import (
    CustomUser,
    Pointage,
    Scan,
    DemandeModification,
    AnomaliePointage,
    AnomalieTraitement,
    PointageAudit,
)


def _is_rh(user):
    return bool(
        user
        and user.is_authenticated
        and (user.is_superuser or getattr(user, "role", None) == "admin")
    )


def _rh_module_permission(self, request):
    return _is_rh(request.user)


def _rh_view_permission(self, request, obj=None):
    return _is_rh(request.user)


def _rh_add_permission(self, request):
    return _is_rh(request.user)


def _rh_change_permission(self, request, obj=None):
    return _is_rh(request.user)


def _rh_delete_permission(self, request, obj=None):
    return _is_rh(request.user)


def _immutable_readonly_fields(self, request, obj=None):
    """Conserve les champs readonly existants et verrouille tous les champs DB."""
    existing = getattr(self, "readonly_fields", ()) or ()
    concrete = tuple(field.name for field in self.model._meta.concrete_fields)
    return tuple(dict.fromkeys((*existing, *concrete)))


def _no_add(self, request):
    return False


def _no_delete(self, request, obj=None):
    return False


def _remove_unsafe_pointage_actions(self):
    """Retire les actions qui pouvaient falsifier directement un pointage."""
    self.actions = [
        action
        for action in (getattr(self, "actions", []) or [])
        if action not in {
            "marquer_present",
            "marquer_retard",
            "marquer_absent",
            "supprimer_selection",
        }
    ]


def _custom_user_fieldsets(self, request, obj=None):
    """Un RH ne peut pas attribuer superuser, groupes ou permissions système."""
    return (
        ("Connexion", {"fields": ("username", "password")}),
        ("Identité", {"fields": ("first_name", "last_name", "email")}),
        ("Rôle & accès", {"fields": ("role", "is_active", "is_staff")}),
    )


def _custom_user_add_fieldsets(self, request):
    return (
        (None, {"classes": ("wide",), "fields": ("username", "password1", "password2")}),
        ("Identité", {"fields": ("first_name", "last_name", "email")}),
        ("Rôle & accès", {"fields": ("role", "is_active")}),
    )


def _custom_user_readonly_fields(self, request, obj=None):
    return ("is_staff",)


def _sanitize_overtime_export(response):
    """Empêche un export de présenter comme H.Supp des heures non autorisées.

    Certains exports historiques recalculent localement `heures_travaillees - 8h`
    au lieu de lire le champ protégé `heures_supplementaires`. La base garantit
    déjà que le champ payable vaut zéro tant que l'autorisation RH est absente,
    mais cette fonction protège aussi les classeurs générés par ces anciens
    calculs sans modifier les heures réellement travaillées.
    """
    if getattr(response, "status_code", 200) != 200:
        return response
    content_type = response.get("Content-Type", "")
    if "spreadsheetml" not in content_type:
        return response

    try:
        workbook = load_workbook(io.BytesIO(response.content))
        worksheet = workbook["Résumé Pointages"] if "Résumé Pointages" in workbook.sheetnames else workbook.active

        if worksheet.max_row < 2 or worksheet.cell(row=2, column=1).value != "Employé":
            return response

        # En-têtes de dates : B2, C2, ... sont du type "Lundi\n03/09/2026".
        day_columns = {}
        for col in range(2, worksheet.max_column):
            value = worksheet.cell(row=2, column=col).value
            if not isinstance(value, str) or "\n" not in value:
                continue
            raw_date = value.split("\n")[-1].strip()
            try:
                from datetime import datetime
                day_columns[col] = datetime.strptime(raw_date, "%d/%m/%Y").date()
            except ValueError:
                continue

        if not day_columns:
            return response

        from .models import Employe

        for row in range(3, worksheet.max_row + 1, 8):
            employee_cell = worksheet.cell(row=row, column=1).value
            if not employee_cell or not isinstance(employee_cell, str):
                continue
            lines = employee_cell.split("\n")
            matricule = lines[-1].strip() if lines else ""
            if not matricule:
                continue
            employe = Employe.objects.filter(matricule=matricule).first()
            if not employe:
                continue

            total_payable = 0
            for col, jour in day_columns.items():
                pointages = Pointage.objects.filter(
                    employe=employe,
                    date_pointage=jour,
                    periode="apres_midi",
                    heures_supplementaires_autorisees=True,
                )
                payable = pointages.first()
                if payable and payable.heures_supplementaires:
                    total_payable += payable.heures_supplementaires.total_seconds()
                    continue

                # Ligne base + 6 = cellule H.Supp du jour dans les deux
                # générateurs de résumé actuels.
                cell = worksheet.cell(row=row + 6, column=col)
                if isinstance(cell.value, str) and cell.value.startswith("H.sup"):
                    cell.value = "—"

            # Ligne de total : ne conserver que l'overtime explicitement validé.
            total_cell = worksheet.cell(row=row, column=worksheet.max_column)
            if isinstance(total_cell.value, str) and "H. Supp :" in total_cell.value:
                lines = total_cell.value.split("\n")
                for i, line in enumerate(lines):
                    if line.strip() == "H. Supp :":
                        if total_payable > 0:
                            hours = int(total_payable // 3600)
                            minutes = int((total_payable % 3600) // 60)
                            formatted = f"{hours}h{minutes:02d}"
                        else:
                            formatted = "0h00"
                        if i + 1 < len(lines):
                            lines[i + 1] = formatted
                        break
                total_cell.value = "\n".join(lines)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        response.content = output.getvalue()
        return response
    except Exception:
        # L'export original reste disponible si un classeur inattendu ne peut
        # pas être analysé ; aucune donnée DB n'est modifiée par ce garde-fou.
        return response


def _wrap_web_export():
    """Patch l'export Web historique avec le contrôle d'autorisation RH."""
    from . import views
    original = getattr(views, "export_resume_excel", None)
    if not original or getattr(original, "_overtime_sanitized", False):
        return

    def wrapped(request, *args, **kwargs):
        response = original(request, *args, **kwargs)
        return _sanitize_overtime_export(response)

    wrapped._overtime_sanitized = True
    views.export_resume_excel = wrapped


@transaction.atomic
def _safe_request_action(self, request, pk, approve):
    """GET affiche une confirmation ; seul POST + CSRF change l'état."""
    demande = get_object_or_404(DemandeModification, pk=pk)

    if not _is_rh(request.user):
        self.message_user(request, "Permission refusée : validation RH requise.", level=messages.ERROR)
        return redirect("admin:index")

    action = "approuver" if approve else "refuser"
    if demande.statut != "en_attente":
        self.message_user(request, f"La demande #{pk} est déjà traitée.", level=messages.WARNING)
        return redirect("../../")

    if request.method == "GET":
        return render(request, "admin/pointage/demande/confirm_action.html", {
            "title": f"Confirmer : {action} la demande #{pk}",
            "demande": demande,
            "action": action,
            "opts": self.model._meta,
            "cancel_url": "../../",
        })

    if request.method != "POST":
        self.message_user(request, "Méthode HTTP non autorisée.", level=messages.ERROR)
        return redirect("../../")

    try:
        if approve:
            self._appliquer_demande(demande)
            demande.statut = "approuvee"
        else:
            demande.statut = "refusee"
        demande.traitee_par = request.user
        demande.date_traitement = timezone.now()
        demande.save(update_fields=("statut", "traitee_par", "date_traitement"))
        self.message_user(
            request,
            f"Demande #{pk} {'approuvée et appliquée' if approve else 'refusée'}.",
            level=messages.SUCCESS,
        )
    except Exception:
        transaction.set_rollback(True)
        self.message_user(
            request,
            "Le traitement a échoué. Aucune modification n'a été enregistrée.",
            level=messages.ERROR,
        )

    return redirect("../../")


def _safe_approve(self, request, pk):
    return _safe_request_action(self, request, pk, True)


def _safe_refuse(self, request, pk):
    return _safe_request_action(self, request, pk, False)


def install():
    registry = admin.site._registry

    if Pointage not in registry:
        from . import admin as _pointage_admin  # noqa: F401
        registry = admin.site._registry

    protected_models = (
        Pointage,
        Scan,
        DemandeModification,
        AnomaliePointage,
        AnomalieTraitement,
        PointageAudit,
    )
    for model in protected_models:
        model_admin = registry.get(model)
        if not model_admin:
            continue
        cls = model_admin.__class__
        cls.has_module_permission = _rh_module_permission
        cls.has_view_permission = _rh_view_permission
        cls.has_add_permission = _rh_add_permission
        cls.has_change_permission = _rh_change_permission
        cls.has_delete_permission = _rh_delete_permission

    for model in (Pointage, Scan, AnomalieTraitement, PointageAudit):
        model_admin = registry.get(model)
        if not model_admin:
            continue
        cls = model_admin.__class__
        cls.get_readonly_fields = _immutable_readonly_fields
        cls.has_add_permission = _no_add
        cls.has_delete_permission = _no_delete

    pointage_admin = registry.get(Pointage)
    if pointage_admin:
        _remove_unsafe_pointage_actions(pointage_admin)
        original_export = getattr(pointage_admin.__class__, "export_excel", None)
        if original_export and not getattr(original_export, "_overtime_sanitized", False):
            def export_excel(self, request, queryset):
                response = original_export(self, request, queryset)
                return _sanitize_overtime_export(response)
            export_excel._overtime_sanitized = True
            pointage_admin.__class__.export_excel = export_excel

    anomalie_admin = registry.get(AnomaliePointage)
    if anomalie_admin:
        cls = anomalie_admin.__class__
        cls.has_add_permission = _no_add
        cls.has_delete_permission = _no_delete
        cls.get_readonly_fields = _immutable_readonly_fields

    demande_admin = registry.get(DemandeModification)
    if demande_admin:
        cls = demande_admin.__class__
        cls.has_add_permission = _no_add
        cls.has_delete_permission = _no_delete
        cls.approuver_view = _safe_approve
        cls.refuser_view = _safe_refuse

    user_admin = registry.get(CustomUser)
    if user_admin:
        cls = user_admin.__class__
        cls.has_module_permission = _rh_module_permission
        cls.has_view_permission = _rh_view_permission
        cls.has_add_permission = _rh_add_permission
        cls.has_change_permission = _rh_change_permission
        cls.has_delete_permission = _rh_delete_permission
        cls.get_fieldsets = _custom_user_fieldsets
        cls.get_add_fieldsets = _custom_user_add_fieldsets
        cls.get_readonly_fields = _custom_user_readonly_fields

        original_save_model = getattr(cls, "save_model", None)

        def save_model(self, request, obj, form, change):
            if not _is_rh(request.user):
                return
            obj.is_superuser = False
            obj.is_staff = getattr(obj, "role", None) == "admin"
            if original_save_model:
                original_save_model(self, request, obj, form, change)

        cls.save_model = save_model

    _wrap_web_export()


install()
