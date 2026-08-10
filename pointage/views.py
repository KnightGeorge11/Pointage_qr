# pointage/views.py

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.generic import ListView, DetailView, DeleteView
from django.urls import reverse_lazy
from django.utils import timezone
from datetime import datetime, timedelta, time
from django.db.models import Q, Sum, Count
from django.http import JsonResponse
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from rest_framework import viewsets, status as drf_status
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated, IsAdminUser
from rest_framework.views import APIView
from .models import Employe, Site, Pointage, Scan, Poste, DemandeModification, AnomaliePointage
from .serializers import (
    EmployeSerializer, SiteSerializer,
    PointageSerializer, PointageDetailSerializer,
    AnomaliePointageSerializer, AnomaliePointageDetailSerializer,
)
from .forms import EmployeForm, SiteForm, PointageForm, PosteForm
from .mixins import DemandeRequiredMixin, AdminCodeRequiredMixin, AdminCodeRequiredForGetMixin
import json
from decimal import Decimal
from .services import process_scan, parse_qr_data
from .anomalies import enregistrer_anomalie, marquer_traitee, marquer_cloturee, compter_anomalies_ouvertes


# ============================================================
# VUES POUR LES DEMANDES DE MODIFICATION (Accepter/Refuser)
# ============================================================

@login_required
def approuver_demande_view(request, pk):
    """Approuve une demande de modification et applique les changements"""
    from django.shortcuts import get_object_or_404
    from .models import DemandeModification
    from .admin import DemandeModificationAdmin
    
    if not request.user.is_staff:
        messages.error(request, "❌ Seul un administrateur peut approuver une demande.")
        return redirect('admin:pointage_demandemodification_changelist')
    
    demande = get_object_or_404(DemandeModification, pk=pk)
    
    if demande.statut != 'en_attente':
        messages.warning(request, f"⚠️ Cette demande a déjà été traitée.")
        return redirect('admin:pointage_demandemodification_changelist')
    
    try:
        admin_instance = DemandeModificationAdmin(DemandeModification, None)
        admin_instance._appliquer_demande(demande)
        
        demande.statut = 'approuvee'
        demande.traitee_par = request.user
        demande.date_traitement = timezone.now()
        demande.save()
        
        messages.success(request, f"✅ Demande #{pk} approuvée et appliquée avec succès.")
    except Exception as e:
        messages.error(request, f"❌ Erreur lors de l'application : {e}")
    
    return redirect('admin:pointage_demandemodification_changelist')


@login_required
def refuser_demande_view(request, pk):
    """Refuse une demande de modification"""
    from django.shortcuts import get_object_or_404
    from .models import DemandeModification
    
    if not request.user.is_staff:
        messages.error(request, "❌ Seul un administrateur peut refuser une demande.")
        return redirect('admin:pointage_demandemodification_changelist')
    
    demande = get_object_or_404(DemandeModification, pk=pk)
    
    if demande.statut != 'en_attente':
        messages.warning(request, f"⚠️ Cette demande a déjà été traitée.")
        return redirect('admin:pointage_demandemodification_changelist')
    
    demande.statut = 'refusee'
    demande.traitee_par = request.user
    demande.date_traitement = timezone.now()
    demande.save()
    
    messages.success(request, f"❌ Demande #{pk} refusée.")
    return redirect('admin:pointage_demandemodification_changelist')


# ---------------------------
# FONCTIONS UTILITAIRES
# ---------------------------

def get_periode_courante():
    now_local = timezone.localtime(timezone.now())
    heure_courante = now_local.time()
    seuil_apres_midi = time(12, 0)
    return 'apres_midi' if heure_courante >= seuil_apres_midi else 'matin'

def get_pointage_du_jour(employe, date_courante):
    return Pointage.objects.filter(
        employe=employe,
        date_pointage=date_courante
    ).order_by('periode')

def get_statut_employe_journee(employe, date_courante):
    pointages = get_pointage_du_jour(employe, date_courante)
    statut = {
        'matin':      {'present': False, 'heure_arrivee': None, 'heure_depart': None, 'site': None},
        'apres_midi': {'present': False, 'heure_arrivee': None, 'heure_depart': None, 'site': None},
        'nuit':       {'present': False, 'heure_arrivee': None, 'heure_depart': None, 'site': None, 'type_journee': None},
    }
    for pointage in pointages:
        if pointage.periode == 'matin':
            statut['matin'].update({'present': True, 'heure_arrivee': pointage.heure_arrivee, 'heure_depart': pointage.heure_depart, 'site': pointage.site})
        elif pointage.periode == 'apres_midi':
            statut['apres_midi'].update({'present': True, 'heure_arrivee': pointage.heure_arrivee, 'heure_depart': pointage.heure_depart, 'site': pointage.site})
        elif pointage.periode == 'nuit':
            statut['nuit'].update({'present': True, 'heure_arrivee': pointage.heure_arrivee, 'heure_depart': pointage.heure_depart, 'site': pointage.site, 'type_journee': pointage.type_journee})
    return statut

def calculer_heures_total_journee(employe, date_courante):
    total = timedelta()
    for pointage in Pointage.objects.filter(employe=employe, date_pointage=date_courante):
        if pointage.heures_travaillees:
            total += pointage.heures_travaillees
    return total


# ---------------------------
# VUES WEB PRINCIPALES
# ---------------------------

@login_required
def dashboard(request):
    today = timezone.localtime(timezone.now()).date()

    total_employes = Employe.objects.filter(actif=True).count()

    today_pointages = Pointage.objects.filter(date_pointage=today)
    presents_aujourdhui = today_pointages.values('employe').distinct().count()
    retards = today_pointages.filter(
        periode__in=['matin', 'apres_midi'], retard__gt=timedelta(0)
    ).count()
    gardes_en_cours = today_pointages.filter(
        periode='nuit', type_journee='garde', heure_depart__isnull=True
    ).count()

    pointages_recents = today_pointages.select_related('employe', 'site').order_by('-date_creation')[:10]
    demandes_en_attente = DemandeModification.objects.filter(statut='en_attente').count()
    anomalies_ouvertes = compter_anomalies_ouvertes()

    week_ago = today - timedelta(days=6)
    daily_stats = Pointage.objects.filter(
        date_pointage__gte=week_ago, date_pointage__lte=today
    ).values('date_pointage').annotate(
        presents=Count('employe', distinct=True),
        retards=Count('id', filter=Q(periode__in=['matin', 'apres_midi'], retard__gt=timedelta(0)))
    )
    stats_by_date = {s['date_pointage']: s for s in daily_stats}

    jours_labels, jours_presents, jours_absents, jours_retards = [], [], [], []
    for i in range(6, -1, -1):
        jour = today - timedelta(days=i)
        jours_labels.append(jour.strftime('%a'))
        s = stats_by_date.get(jour, {})
        p = s.get('presents', 0)
        r = s.get('retards', 0)
        jours_presents.append(p)
        jours_absents.append(total_employes - p)
        jours_retards.append(r)

    four_weeks_ago = today - timedelta(days=today.weekday() + 21)
    all_weekly = Pointage.objects.filter(
        date_pointage__gte=four_weeks_ago, date_pointage__lte=today
    )

    semaines_labels, semaines_taux_presence, semaines_taux_punctualite = [], [], []
    for i in range(3, -1, -1):
        start_of_week = today - timedelta(days=today.weekday() + 7*i)
        end_of_week   = start_of_week + timedelta(days=6)
        semaines_labels.append(f"Sem. {start_of_week.strftime('%d/%m')}")
        semaine_qs = all_weekly.filter(date_pointage__gte=start_of_week, date_pointage__lte=end_of_week)
        employes_semaine = semaine_qs.values('employe').distinct().count()
        taux_presence = (employes_semaine / total_employes * 100) if total_employes else 0
        pointages_sans_ret = semaine_qs.filter(periode__in=['matin', 'apres_midi'], retard=timedelta(0)).count()
        total_pointages = semaine_qs.filter(periode__in=['matin', 'apres_midi']).count()
        taux_punctualite = (pointages_sans_ret / total_pointages * 100) if total_pointages else 0
        semaines_taux_presence.append(round(taux_presence, 1))
        semaines_taux_punctualite.append(round(taux_punctualite, 1))

    postes_qs = Poste.objects.annotate(
        employes_count=Count('employes', filter=Q(employes__actif=True))
    ).filter(employes_count__gt=0)
    postes_stats = [{'nom': p.nom, 'count': p.employes_count, 'couleur': p.couleur} for p in postes_qs]

    context = {
        'total_employes':       total_employes,
        'presents_aujourdhui':  presents_aujourdhui,
        'absents_aujourdhui':   total_employes - presents_aujourdhui,
        'retards_aujourdhui':   retards,
        'gardes_en_cours':      gardes_en_cours,
        'pointages_recents':    pointages_recents,
        'demandes_en_attente':   demandes_en_attente,
        'anomalies_ouvertes':    anomalies_ouvertes,
        'aujourdhui':           today,
        'daily_data':    {'presents': presents_aujourdhui, 'absents': total_employes - presents_aujourdhui, 'retards': retards, 'gardes': gardes_en_cours},
        'weekly_data':   {'labels': jours_labels, 'presents': jours_presents, 'absents': jours_absents, 'retards': jours_retards},
        'evolution_data':{'labels': semaines_labels, 'taux_presence': semaines_taux_presence, 'taux_punctualite': semaines_taux_punctualite},
        'postes_data':   postes_stats,
    }
    return render(request, 'pointage/dashboard.html', context)


@login_required
def index(request):
    today = timezone.localtime(timezone.now()).date()
    pointages_aujourdhui = Pointage.objects.filter(date_pointage=today)
    total_employes       = Employe.objects.filter(actif=True).count()
    presents_aujourdhui  = pointages_aujourdhui.values('employe').distinct().count()
    retards              = pointages_aujourdhui.filter(periode__in=['matin', 'apres_midi'], retard__gt=timedelta(0))
    pointages_recents    = Pointage.objects.filter(date_pointage=today).select_related('employe', 'site').order_by('-date_creation')[:10]
    context = {
        'total_employes':      total_employes,
        'presents_aujourdhui': presents_aujourdhui,
        'absents_aujourdhui':  total_employes - presents_aujourdhui,
        'retards_aujourdhui':  retards.count(),
        'pointages_recents':   pointages_recents,
        'aujourdhui':          today,
    }
    return render(request, 'pointage/index.html', context)


@login_required
def scanner_view(request):
    sites = Site.objects.all()

    if request.method == 'POST':
        raw_qr      = request.POST.get('qr_data', '').strip()
        matricule   = request.POST.get('matricule', '').strip()
        site_id     = request.POST.get('site_id')
        periode_type = request.POST.get('periode_type', 'auto')

        if raw_qr:
            parsed = parse_qr_data(raw_qr)
            if not parsed:
                messages.error(request, "❌ Format QR invalide.")
                return redirect('scanner')
            mat   = parsed['matricule']
            token = parsed['token']
        elif matricule:
            try:
                emp   = Employe.objects.get(matricule=matricule, actif=True)
                mat   = emp.matricule
                token = str(emp.qr_code_token)
            except Employe.DoesNotExist:
                messages.error(request, f"❌ Employé {matricule} non trouvé.")
                return redirect('scanner')
        else:
            messages.error(request, "❌ QR code ou matricule requis.")
            return redirect('scanner')

        if not site_id:
            messages.error(request, "❌ Veuillez sélectionner un site.")
            return redirect('scanner')

        mode = 'garde' if periode_type == 'garde' else 'auto'

        result = process_scan(
            matricule=mat,
            qr_token=token,
            site_id=int(site_id),
            mode=mode
        )

        if result['status'] == 'success':
            messages.success(request, f"✅ {result['message']}")
        elif result['status'] == 'warning':
            messages.warning(request, f"⚠️ {result['message']}")
        else:
            messages.error(request, f"❌ {result['message']}")

        return redirect('scanner')

    today                = timezone.localtime(timezone.now()).date()
    total_employes       = Employe.objects.filter(actif=True).count()
    pointages_aujourdhui = Pointage.objects.filter(date_pointage=today)
    presents_aujourdhui  = pointages_aujourdhui.values('employe').distinct().count()
    gardes_en_cours      = Pointage.objects.filter(
        date_pointage=today, periode='nuit',
        type_journee='garde', heure_depart__isnull=True
    ).count()

    context = {
        'sites':               sites,
        'aujourdhui':          today,
        'total_employes':      total_employes,
        'presents_aujourdhui': presents_aujourdhui,
        'gardes_en_cours':     gardes_en_cours,
        'periode_actuelle':    get_periode_courante().capitalize(),
    }
    return render(request, 'pointage/scanner.html', context)


# -------------------------------------------------------
# HELPER — Crée une demande et redirige avec message
# -------------------------------------------------------

def _creer_demande(request, type_action, cible, form=None, cible_id=None):
    """Sérialise le formulaire et crée une DemandeModification en attente."""
    donnees = {}
    if form:
        for field, value in form.cleaned_data.items():
            if hasattr(value, 'pk'):
                donnees[field] = value.pk
            elif hasattr(value, '__iter__') and not isinstance(value, str):
                donnees[field] = [v.pk if hasattr(v, 'pk') else v for v in value]
            else:
                donnees[field] = str(value) if value is not None else None

    DemandeModification.objects.create(
        demandeur=request.user,
        type_action=type_action,
        cible=cible,
        cible_id=cible_id,
        donnees=donnees,
    )


# ---------------------------
# VUES POUR LES EMPLOYÉS
# ---------------------------

class EmployeListView(LoginRequiredMixin, ListView):
    model               = Employe
    template_name       = 'pointage/employes.html'
    context_object_name = 'employes'

    def get_queryset(self):
        return Employe.objects.filter(actif=True).select_related('poste').order_by('matricule')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['demandes_en_attente'] = DemandeModification.objects.filter(
            statut='en_attente', cible='employe'
        ).count()
        return context


@login_required
def employe_create_view(request):
    if request.method == 'POST':
        form = EmployeForm(request.POST)
        if form.is_valid():
            if request.user.is_staff:
                employe = form.save()
                messages.success(request, f"✅ Employé {employe.get_nom_complet()} créé avec succès.")
                return redirect('employes')
            else:
                _creer_demande(request, 'create', 'employe', form=form)
                messages.success(request, "✅ Votre demande d'ajout d'employé a été envoyée à l'administrateur.")
                return redirect('employes')
    else:
        form = EmployeForm()
    return render(request, 'pointage/employe_form.html', {
        'form': form,
        'mode': 'demande' if not request.user.is_staff else 'creation',
        'titre': 'Ajout d\'employé' if request.user.is_staff else 'Demande d\'ajout d\'employé',
    })


@login_required
def employe_update_view(request, pk):
    employe = get_object_or_404(Employe, pk=pk)
    if request.method == 'POST':
        form = EmployeForm(request.POST, instance=employe)
        if form.is_valid():
            if request.user.is_staff:
                form.save()
                messages.success(request, f"✅ Employé {employe.get_nom_complet()} modifié avec succès.")
                return redirect('employes')
            else:
                _creer_demande(request, 'update', 'employe', form=form, cible_id=pk)
                messages.success(request, "✅ Votre demande de modification a été envoyée à l'administrateur.")
                return redirect('employes')
    else:
        form = EmployeForm(instance=employe)
    return render(request, 'pointage/employe_form.html', {
        'form': form,
        'object': employe,
        'mode': 'demande' if not request.user.is_staff else 'modification',
        'titre': f'Modification de {employe.get_nom_complet()}' if request.user.is_staff else f'Demande de modification — {employe.get_nom_complet()}',
    })


@login_required
def employe_delete_view(request, pk):
    """
    Suppression d'un employé.
    - Admin (is_staff) : supprime directement avec confirmation
    - Utilisateur normal : crée une demande de suppression
    """
    employe = get_object_or_404(Employe, pk=pk)
    
    if request.user.is_staff:
        # ADMIN : Suppression directe
        if request.method == 'POST':
            nom = employe.get_nom_complet()
            employe.delete()
            messages.success(request, f"✅ Employé {nom} supprimé avec succès.")
            return redirect('employes')
        # GET : Afficher la confirmation
        return render(request, 'pointage/employe_confirm_delete.html', {
            'object': employe,
            'mode': 'suppression',
        })
    else:
        # UTILISATEUR NORMAL : Demande de suppression
        if request.method == 'POST':
            _creer_demande(request, 'delete', 'employe', cible_id=pk)
            messages.success(request, "✅ Votre demande de suppression a été envoyée à l'administrateur.")
            return redirect('employes')
        # GET : Afficher la confirmation de demande
        return render(request, 'pointage/employe_confirm_delete.html', {
            'object': employe,
            'mode': 'demande',
        })


# ---------------------------
# VUES POUR LES SITES
# ---------------------------

class SiteListView(LoginRequiredMixin, ListView):
    model               = Site
    template_name       = 'pointage/sites.html'
    context_object_name = 'sites'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['demandes_en_attente'] = DemandeModification.objects.filter(
            statut='en_attente', cible='site'
        ).count()
        return context


@login_required
def site_create_view(request):
    if request.method == 'POST':
        form = SiteForm(request.POST)
        if form.is_valid():
            if request.user.is_staff:
                site = form.save()
                messages.success(request, f"✅ Site {site.nom} créé avec succès.")
                return redirect('sites')
            else:
                _creer_demande(request, 'create', 'site', form=form)
                messages.success(request, "✅ Votre demande d'ajout de site a été envoyée à l'administrateur.")
                return redirect('sites')
    else:
        form = SiteForm()
    return render(request, 'pointage/site_form.html', {
        'form': form,
        'mode': 'demande' if not request.user.is_staff else 'creation',
        'titre': 'Ajout de site' if request.user.is_staff else 'Demande d\'ajout de site',
    })


@login_required
def site_update_view(request, pk):
    site = get_object_or_404(Site, pk=pk)
    if request.method == 'POST':
        form = SiteForm(request.POST, instance=site)
        if form.is_valid():
            if request.user.is_staff:
                form.save()
                messages.success(request, f"✅ Site {site.nom} modifié avec succès.")
                return redirect('sites')
            else:
                _creer_demande(request, 'update', 'site', form=form, cible_id=pk)
                messages.success(request, "✅ Votre demande de modification a été envoyée à l'administrateur.")
                return redirect('sites')
    else:
        form = SiteForm(instance=site)
    return render(request, 'pointage/site_form.html', {
        'form': form,
        'object': site,
        'mode': 'demande' if not request.user.is_staff else 'modification',
        'titre': f'Modification de {site.nom}' if request.user.is_staff else f'Demande de modification — {site.nom}',
    })


@login_required
def site_delete_view(request, pk):
    """
    Suppression d'un site.
    - Admin (is_staff) : supprime directement avec confirmation
    - Utilisateur normal : crée une demande de suppression
    """
    site = get_object_or_404(Site, pk=pk)
    
    if request.user.is_staff:
        # ADMIN : Suppression directe
        if request.method == 'POST':
            nom = site.nom
            site.delete()
            messages.success(request, f"✅ Site {nom} supprimé avec succès.")
            return redirect('sites')
        # GET : Afficher la confirmation
        return render(request, 'pointage/site_confirm_delete.html', {
            'object': site,
            'mode': 'suppression',
        })
    else:
        # UTILISATEUR NORMAL : Demande de suppression
        if request.method == 'POST':
            _creer_demande(request, 'delete', 'site', cible_id=pk)
            messages.success(request, "✅ Votre demande de suppression a été envoyée à l'administrateur.")
            return redirect('sites')
        # GET : Afficher la confirmation de demande
        return render(request, 'pointage/site_confirm_delete.html', {
            'object': site,
            'mode': 'demande',
        })


# ---------------------------
# VUES POUR LES POSTES
# ---------------------------

class PosteListView(LoginRequiredMixin, ListView):
    model               = Poste
    template_name       = 'pointage/postes.html'
    context_object_name = 'postes'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['demandes_en_attente'] = DemandeModification.objects.filter(
            statut='en_attente', cible='poste'
        ).count()
        return context


@login_required
def poste_create_view(request):
    if request.method == 'POST':
        form = PosteForm(request.POST)
        if form.is_valid():
            if request.user.is_staff:
                poste = form.save()
                messages.success(request, f"✅ Poste {poste.nom} créé avec succès.")
                return redirect('postes')
            else:
                _creer_demande(request, 'create', 'poste', form=form)
                messages.success(request, "✅ Votre demande d'ajout de poste a été envoyée à l'administrateur.")
                return redirect('postes')
    else:
        form = PosteForm()
    return render(request, 'pointage/poste_form.html', {
        'form': form,
        'mode': 'demande' if not request.user.is_staff else 'creation',
        'titre': 'Ajout de poste' if request.user.is_staff else 'Demande d\'ajout de poste',
    })


@login_required
def poste_update_view(request, pk):
    poste = get_object_or_404(Poste, pk=pk)
    if request.method == 'POST':
        form = PosteForm(request.POST, instance=poste)
        if form.is_valid():
            if request.user.is_staff:
                form.save()
                messages.success(request, f"✅ Poste {poste.nom} modifié avec succès.")
                return redirect('postes')
            else:
                _creer_demande(request, 'update', 'poste', form=form, cible_id=pk)
                messages.success(request, "✅ Votre demande de modification a été envoyée à l'administrateur.")
                return redirect('postes')
    else:
        form = PosteForm(instance=poste)
    return render(request, 'pointage/poste_form.html', {
        'form': form,
        'object': poste,
        'mode': 'demande' if not request.user.is_staff else 'modification',
        'titre': f'Modification de {poste.nom}' if request.user.is_staff else f'Demande de modification — {poste.nom}',
    })


@login_required
def poste_delete_view(request, pk):
    """
    Suppression d'un poste.
    - Admin (is_staff) : supprime directement avec confirmation
    - Utilisateur normal : crée une demande de suppression
    """
    poste = get_object_or_404(Poste, pk=pk)
    
    if request.user.is_staff:
        # ADMIN : Suppression directe
        if request.method == 'POST':
            nom = poste.nom
            poste.delete()
            messages.success(request, f"✅ Poste {nom} supprimé avec succès.")
            return redirect('postes')
        # GET : Afficher la confirmation
        return render(request, 'pointage/poste_confirm_delete.html', {
            'object': poste,
            'mode': 'suppression',
        })
    else:
        # UTILISATEUR NORMAL : Demande de suppression
        if request.method == 'POST':
            _creer_demande(request, 'delete', 'poste', cible_id=pk)
            messages.success(request, "✅ Votre demande de suppression a été envoyée à l'administrateur.")
            return redirect('postes')
        # GET : Afficher la confirmation de demande
        return render(request, 'pointage/poste_confirm_delete.html', {
            'object': poste,
            'mode': 'demande',
        })


# ---------------------------
# VUES POUR LES POINTAGES
# ---------------------------

class PointageListView(LoginRequiredMixin, ListView):
    model               = Pointage
    template_name       = 'pointage/pointages.html'
    context_object_name = 'jours'
    paginate_by         = 20

    def get_queryset(self):
        queryset = Pointage.objects.all().select_related('employe', 'site').prefetch_related('scans')

        date_debut = self.request.GET.get('date_debut')
        date_fin   = self.request.GET.get('date_fin')
        if date_debut:
            queryset = queryset.filter(date_pointage__gte=date_debut)
        if date_fin:
            queryset = queryset.filter(date_pointage__lte=date_fin)

        employe_filter = self.request.GET.get('employe')
        if employe_filter:
            queryset = queryset.filter(employe_id=employe_filter)

        site_filter = self.request.GET.get('site')
        if site_filter:
            queryset = queryset.filter(site_id=site_filter)

        periode_type = self.request.GET.get('periode_type')
        if periode_type == 'jour':
            queryset = queryset.filter(type_journee='normal')
        elif periode_type == 'nuit':
            queryset = queryset.filter(type_journee='garde')

        return queryset.order_by('-date_pointage', 'employe__nom')

    def get_context_data(self, **kwargs):
        context  = super().get_context_data(**kwargs)
        queryset = self.get_queryset()

        jours_dict = {}
        for pointage in queryset:
            key = (pointage.employe_id, pointage.date_pointage)
            if key not in jours_dict:
                jours_dict[key] = {
                    'date': pointage.date_pointage, 'employe': pointage.employe, 'site': pointage.site,
                    'matin': None, 'apres_midi': None, 'nuit': None,
                    'scans': [], 'scan_map': {},
                }
            if pointage.periode == 'matin':
                jours_dict[key]['matin'] = pointage
            elif pointage.periode == 'apres_midi':
                jours_dict[key]['apres_midi'] = pointage
            elif pointage.periode == 'nuit':
                jours_dict[key]['nuit'] = pointage

        jours_list = []
        for key, jour in jours_dict.items():
            scan_map = {}
            for periode_key in ['matin', 'apres_midi', 'nuit']:
                p = jour[periode_key]
                if p:
                    for scan in p.scans.all().order_by('timestamp'):
                        scan_map[scan.type_scan] = {
                            'heure': timezone.localtime(scan.timestamp),
                            'site':  scan.site.nom if scan.site else '',
                        }
            jour['scan_map'] = scan_map

            heures_total = timedelta()
            retard_total = timedelta()
            if jour['matin']:
                heures_total += jour['matin'].heures_travaillees or timedelta()
                retard_total += jour['matin'].retard or timedelta()
            if jour['apres_midi']:
                heures_total += jour['apres_midi'].heures_travaillees or timedelta()
                retard_total += jour['apres_midi'].retard or timedelta()
            if jour['nuit']:
                heures_total += jour['nuit'].heures_travaillees or timedelta()

            jour['heures_total']  = heures_total
            jour['retard_total']  = retard_total
            jour['heures_sup']    = max(heures_total - timedelta(hours=8), timedelta())
            jour['statut_global'] = 'present' if (jour['matin'] or jour['apres_midi'] or jour['nuit']) else 'absent'
            jour['badge_type']    = 'garde' if jour['nuit'] and jour['nuit'].type_journee == 'garde' else 'normal'
            jours_list.append(jour)

        jours_list.sort(key=lambda x: x['date'], reverse=True)

        total_heures    = timedelta()
        total_retard    = timedelta()
        unique_employes = set()
        for jour in jours_list:
            total_heures  += jour['heures_total']
            total_retard  += jour['retard_total']
            unique_employes.add(jour['employe'].id)

        context['total_heures']          = total_heures
        context['total_retard']          = total_retard
        context['unique_employes_count'] = len(unique_employes)
        context['total_journees']        = len(jours_list)

        paginator = Paginator(jours_list, self.paginate_by)
        page = self.request.GET.get('page')
        try:
            jours_page = paginator.page(page)
        except PageNotAnInteger:
            jours_page = paginator.page(1)
        except EmptyPage:
            jours_page = paginator.page(paginator.num_pages)

        context['jours']             = jours_page
        context['employes']          = Employe.objects.filter(actif=True)
        context['sites']             = Site.objects.all()
        context['filter_date_debut'] = self.request.GET.get('date_debut', '')
        context['filter_date_fin']   = self.request.GET.get('date_fin', '')
        return context


class PointageDetailView(LoginRequiredMixin, DetailView):
    model               = Pointage
    template_name       = 'pointage/pointage_detail.html'
    context_object_name = 'pointage'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['scans'] = self.object.scans.filter(actif=True).order_by('timestamp')
        return context


class PointageDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model         = Pointage
    template_name = 'pointage/pointage_confirm_delete.html'
    success_url   = reverse_lazy('pointages')

    def test_func(self):
        return self.request.user.is_staff

    def handle_no_permission(self):
        messages.error(self.request, "❌ Seul un administrateur peut supprimer un pointage.")
        return redirect('pointages')

    def delete(self, request, *args, **kwargs):
        messages.success(request, "✅ Pointage supprimé avec succès!")
        return super().delete(request, *args, **kwargs)


# ============================================================
# ANOMALIES DE POINTAGE
# ============================================================

@login_required
def alertes_rh_view(request):
    if request.method == 'POST':
        if not request.user.is_staff:
            messages.error(request, "❌ Seul un administrateur ou RH peut traiter une anomalie.")
            return redirect('alertes_rh')
        
        anomalie = get_object_or_404(AnomaliePointage, pk=request.POST.get('anomalie_id'))
        action = request.POST.get('action')
        
        try:
            if action == 'traiter':
                commentaire = request.POST.get('commentaire', '').strip()
                champ = request.POST.get('champ_corrige', '').strip()
                ancienne = request.POST.get('ancienne_valeur', '').strip()
                nouvelle = request.POST.get('nouvelle_valeur', '').strip()
                corrections = []
                if champ:
                    corrections.append({
                        'champ': champ,
                        'ancienne_valeur': ancienne,
                        'nouvelle_valeur': nouvelle,
                    })
                marquer_traitee(anomalie, request.user, commentaire=commentaire, corrections=corrections)
                messages.success(request, f"✅ Anomalie #{anomalie.pk} marquée comme traitée.")
                
            elif action == 'cloturer':
                marquer_cloturee(anomalie, request.user)
                messages.success(request, f"🔒 Anomalie #{anomalie.pk} clôturée.")
                
        except ValueError as e:
            messages.error(request, f"❌ {e}")
        except PermissionError as e:
            messages.error(request, f"❌ {e}")
            
        return redirect('alertes_rh')

    filter_type = request.GET.get('type', '')
    filter_statut = request.GET.get('statut', '')
    filter_search = request.GET.get('search', '').strip()

    qs = AnomaliePointage.objects.select_related(
        'employe', 'site', 'traitement', 'traitement__administrateur', 'cloturee_par'
    )
    
    if filter_type:
        qs = qs.filter(type=filter_type)
    if filter_statut:
        qs = qs.filter(statut=filter_statut)
    if filter_search:
        qs = qs.filter(
            Q(employe__nom__icontains=filter_search) |
            Q(employe__prenom__icontains=filter_search) |
            Q(employe__matricule__icontains=filter_search) |
            Q(matricule_scanne__icontains=filter_search)
        )

    total_alertes = qs.count()
    ouvertes = AnomaliePointage.objects.filter(statut=AnomaliePointage.STATUT_OUVERTE).count()
    traitees = AnomaliePointage.objects.filter(statut=AnomaliePointage.STATUT_TRAITEE).count()
    cloturees = AnomaliePointage.objects.filter(statut=AnomaliePointage.STATUT_CLOTUREE).count()
    non_traitees = ouvertes

    paginator = Paginator(qs, 20)
    alertes = paginator.get_page(request.GET.get('page'))

    context = {
        'alertes': alertes,
        'non_traitees': non_traitees,
        'ouvertes': ouvertes,
        'traitees': traitees,
        'cloturees': cloturees,
        'total_alertes': total_alertes,
        'types_alerte': AnomaliePointage.TYPE_CHOICES,
        'filter_type': filter_type,
        'filter_statut': filter_statut,
        'filter_search': filter_search,
        'user_is_staff': request.user.is_staff,
    }
    return render(request, 'admin/pointage/alerte/alertes_rh.html', context)


@login_required
def export_resume_excel(request):
    """Export résumé par employé entre 2 dates — format tableau par jour"""
    from collections import defaultdict

    date_debut_str = request.GET.get('date_debut')
    date_fin_str   = request.GET.get('date_fin')
    employe_filter = request.GET.get('employe')
    site_filter    = request.GET.get('site')

    try:
        if date_debut_str and date_fin_str:
            date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
            date_fin   = datetime.strptime(date_fin_str,   '%Y-%m-%d').date()
        else:
            date_debut = Pointage.objects.order_by('date_pointage').values_list('date_pointage', flat=True).first()
            date_fin   = Pointage.objects.order_by('-date_pointage').values_list('date_pointage', flat=True).first()
            if not date_debut or not date_fin:
                return HttpResponse("Aucun pointage trouvé en base de données.", status=404)
    except ValueError:
        date_debut = Pointage.objects.order_by('date_pointage').values_list('date_pointage', flat=True).first()
        date_fin   = Pointage.objects.order_by('-date_pointage').values_list('date_pointage', flat=True).first()
        if not date_debut or not date_fin:
            today = timezone.localtime(timezone.now()).date()
            date_debut = date_fin = today

    try:
        def work_days(d1, d2):
            days, d = [], d1
            while d <= d2:
                days.append(d)
                d += timedelta(days=1)
            return days

        days = work_days(date_debut, date_fin)
        JOURS_FR = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']

        qs = Pointage.objects.filter(
            date_pointage__gte=date_debut,
            date_pointage__lte=date_fin
        ).select_related('employe', 'site', 'employe__poste').order_by(
            'employe__nom', 'employe__prenom', 'date_pointage'
        )
        if employe_filter:
            qs = qs.filter(employe_id=employe_filter)
        if site_filter:
            qs = qs.filter(site_id=site_filter)

        emp_data = defaultdict(lambda: defaultdict(lambda: {'matin': None, 'apres_midi': None, 'nuit': None}))
        emp_info = {}
        for p in qs:
            emp_info[p.employe.id] = (p.employe.id, p.employe.get_nom_complet(), p.employe.matricule)
            emp_data[p.employe.id][p.date_pointage][p.periode] = p

        def fmt_duree(pointage):
            if not pointage or not pointage.heures_travaillees:
                return '—'
            return pointage.get_duree_formatee()

        def fmt_time(t):
            return t.strftime('%H:%M') if t else '—'

        BLUE       = '1E3A5F'
        BLUE_LIGHT = 'D6E4F0'
        ORANGE_BG  = 'FEF3C7'
        ORANGE_FG  = 'D97706'
        GREEN_BG   = 'DCFCE7'
        GREEN_FG   = '15803D'
        RED_BG     = 'FEE2E2'
        RED_FG     = 'B91C1C'
        PURPLE_BG  = 'EDE9FE'
        PURPLE_FG  = '7C3AED'
        PURPLE_MED = '4C1D95'
        NIGHT_BG   = '1E1B4B'
        NIGHT_MID  = '2D1F4E'
        NIGHT_FG   = 'A5B4FC'
        DARK       = '1A1A1A'
        GREY_LIGHT = 'F5F5F7'
        GREY_MID   = 'E5E5E5'
        WHITE      = 'FFFFFF'
        TOTAL_BG   = 'EEF2FF'

        def sd(color=GREY_MID, style='thin'):
            return Side(style=style, color=color)

        def b_all(color=GREY_MID):
            s = sd(color)
            return Border(left=s, right=s, top=s, bottom=s)

        def b_outer(color=BLUE):
            s = Side(style='medium', color=color)
            return Border(left=s, right=s, top=s, bottom=s)

        def b_bottom(color=BLUE):
            return Border(left=sd(), right=sd(), top=sd(),
                          bottom=Side(style='medium', color=color))

        def sc(c, value='', bg=WHITE, fg=DARK, bold=False, size=9,
               halign='center', valign='center', wrap=False, border=None, italic=False):
            c.value = value
            c.font  = Font(name='Arial', bold=bold, color=fg, size=size, italic=italic)
            c.fill  = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
            if border:
                c.border = border

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Résumé Pointages"
        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation   = 'landscape'
        ws.page_setup.fitToPage     = True
        ws.page_setup.fitToWidth    = 1

        COL_EMP   = 1
        COL_DAYS  = 2
        COL_TOTAL = COL_DAYS + len(days)

        ws.column_dimensions[get_column_letter(COL_EMP)].width = 15
        for i in range(len(days)):
            ws.column_dimensions[get_column_letter(COL_DAYS + i)].width = 14
        ws.column_dimensions[get_column_letter(COL_TOTAL)].width = 18

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COL_TOTAL)
        sc(ws['A1'],
           value=f"RÉSUMÉ DES POINTAGES  ·  Du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}",
           bg=BLUE, fg=WHITE, bold=True, size=13)
        ws.row_dimensions[1].height = 36

        HEADER_ROW = 2
        sc(ws.cell(row=HEADER_ROW, column=COL_EMP), value='Employé',
           bg=BLUE, fg=WHITE, bold=True, size=9, border=b_all(BLUE))
        for i, d in enumerate(days):
            label = f"{JOURS_FR[d.weekday()]}\n{d.strftime('%d/%m/%Y')}"
            sc(ws.cell(row=HEADER_ROW, column=COL_DAYS + i), value=label,
               bg=BLUE, fg=WHITE, bold=True, size=8, wrap=True, border=b_all(BLUE))
        sc(ws.cell(row=HEADER_ROW, column=COL_TOTAL), value='TOTAL',
           bg=DARK, fg=WHITE, bold=True, size=10, border=b_all(DARK))
        ws.row_dimensions[HEADER_ROW].height = 28

        ROWS_PER_EMP = 8
        SUB_H        = [8, 14, 18, 14, 18, 16, 16, 6]
        START_ROW    = 3

        for emp_idx, (emp_id, day_map) in enumerate(emp_data.items()):
            base = START_ROW + emp_idx * ROWS_PER_EMP
            _, nom_complet, matricule = emp_info[emp_id]

            for i, h in enumerate(SUB_H):
                ws.row_dimensions[base + i].height = h

            ws.merge_cells(start_row=base, start_column=COL_EMP,
                           end_row=base + 6, end_column=COL_EMP)
            sc(ws.cell(row=base, column=COL_EMP),
               value=f"{nom_complet}\n{matricule}",
               bg=GREY_LIGHT, fg=DARK, bold=True, size=9,
               wrap=True, halign='center', valign='center', border=b_outer())
            ws.cell(row=base + 7, column=COL_EMP).fill   = PatternFill('solid', start_color=GREY_MID)
            ws.cell(row=base + 7, column=COL_EMP).border = b_bottom()

            tot_retard = timedelta()
            tot_trav   = timedelta()
            tot_sup    = timedelta()
            tot_gardes = 0

            for i, d in enumerate(days):
                col    = COL_DAYS + i
                pt_map = day_map.get(d, {})
                matin  = pt_map.get('matin')
                apm    = pt_map.get('apres_midi')
                nuit   = pt_map.get('nuit')
                bg_day = WHITE if i % 2 == 0 else 'FAFAFA'

                if nuit and nuit.type_journee == 'garde':
                    tot_gardes += 1
                    h_garde     = nuit.heures_travaillees or timedelta()
                    tot_trav   += h_garde
                    terminee    = bool(nuit.heure_depart)

                    sc(ws.cell(row=base, column=col), bg=NIGHT_MID,
                       border=Border(top=Side(style='medium', color=PURPLE_FG),
                                     left=sd(), right=sd()))
                    ws.merge_cells(start_row=base + 1, start_column=col,
                                   end_row=base + 2,   end_column=col)
                    sc(ws.cell(row=base + 1, column=col),
                       value='🌙 Garde de nuit',
                       bg=PURPLE_BG, fg=PURPLE_FG, bold=True, size=9,
                       wrap=True, halign='center', valign='center',
                       border=b_all(PURPLE_FG))
                    sc(ws.cell(row=base + 3, column=col),
                       value='Début  →  Fin',
                       bg=NIGHT_MID, fg=NIGHT_FG, bold=True, size=8,
                       border=b_all(PURPLE_MED))
                    arr_g = fmt_time(nuit.heure_arrivee)
                    dep_g = fmt_time(nuit.heure_depart)
                    sc(ws.cell(row=base + 4, column=col),
                       value=f"{arr_g}  →  {dep_g}",
                       bg=NIGHT_BG, fg=WHITE, bold=True, size=9,
                       border=b_all(PURPLE_MED))
                    sc(ws.cell(row=base + 5, column=col),
                       value=f"Durée : {fmt_duree(nuit)}",
                       bg=PURPLE_BG, fg=PURPLE_FG, bold=True, size=8,
                       border=b_all(PURPLE_FG))
                    if terminee:
                        sc(ws.cell(row=base + 6, column=col),
                           value='✓ Terminée', bg=GREEN_BG, fg=GREEN_FG,
                           bold=True, size=8, border=b_all())
                    else:
                        sc(ws.cell(row=base + 6, column=col),
                           value='⏳ En cours', bg=ORANGE_BG, fg=ORANGE_FG,
                           bold=True, size=8, border=b_all())
                    sc(ws.cell(row=base + 7, column=col), bg=GREY_LIGHT,
                       border=Border(bottom=Side(style='medium', color=PURPLE_FG),
                                     left=sd(), right=sd()))
                else:
                    has_data = matin or apm
                    h_trav   = timedelta()
                    h_ret    = timedelta()
                    if matin:
                        h_trav += matin.heures_travaillees or timedelta()
                        h_ret  += matin.retard or timedelta()
                    if apm:
                        h_trav += apm.heures_travaillees or timedelta()
                        h_ret  += apm.retard or timedelta()
                    h_sup = max(timedelta(), h_trav - timedelta(hours=8))

                    tot_trav   += h_trav
                    tot_retard += h_ret
                    tot_sup    += h_sup

                    sc(ws.cell(row=base, column=col),
                       bg=bg_day if has_data else GREY_LIGHT,
                       border=Border(top=Side(style='medium', color=BLUE),
                                     left=sd(), right=sd()))
                    sc(ws.cell(row=base + 1, column=col),
                       value='Matin', bg=ORANGE_BG, fg=ORANGE_FG,
                       bold=True, size=8, border=b_all())
                    arr_m = fmt_time(matin.heure_arrivee if matin else None)
                    dep_m = fmt_time(matin.heure_depart  if matin else None)
                    sc(ws.cell(row=base + 2, column=col),
                       value=f"{arr_m}  →  {dep_m}" if has_data else '—',
                       bg=bg_day, fg=DARK, bold=True, size=9, border=b_all())
                    sc(ws.cell(row=base + 3, column=col),
                       value='Après-midi', bg=BLUE_LIGHT, fg=BLUE,
                       bold=True, size=8, border=b_all())
                    arr_s = fmt_time(apm.heure_arrivee if apm else None)
                    dep_s = fmt_time(apm.heure_depart  if apm else None)
                    sc(ws.cell(row=base + 4, column=col),
                       value=f"{arr_s}  →  {dep_s}" if has_data else '—',
                       bg=bg_day, fg=DARK, bold=True, size=9, border=b_all())
                    sc(ws.cell(row=base + 5, column=col),
                       value=f"Retard : {matin.get_retard_minutes if matin else 0}min" if h_ret.total_seconds() > 0 else '—',
                       bg=RED_BG if h_ret.total_seconds() > 0 else bg_day,
                       fg=RED_FG if h_ret.total_seconds() > 0 else '999999',
                       size=8, italic=True, border=b_all())
                    sc(ws.cell(row=base + 6, column=col),
                       value=f"H.sup : {fmt_duree(Pointage(heures_travaillees=h_sup))}" if h_sup.total_seconds() > 0 else '—',
                       bg=GREEN_BG if h_sup.total_seconds() > 0 else bg_day,
                       fg=GREEN_FG if h_sup.total_seconds() > 0 else '999999',
                       size=8, italic=True, border=b_all())
                    sc(ws.cell(row=base + 7, column=col), bg=GREY_LIGHT,
                       border=Border(bottom=Side(style='medium', color=BLUE),
                                     left=sd(), right=sd()))

            ws.merge_cells(start_row=base, start_column=COL_TOTAL,
                           end_row=base + 6, end_column=COL_TOTAL)
            
            total_lines = [
                f"Retards :\n{Pointage(heures_travaillees=tot_retard).get_duree_formatee()}" if tot_retard.total_seconds() > 0 else "Retards :\n0h00",
                f"\nH. Travaillées :\n{Pointage(heures_travaillees=tot_trav).get_duree_formatee()}",
                f"\nH. Supp :\n{Pointage(heures_travaillees=tot_sup).get_duree_formatee()}" if tot_sup.total_seconds() > 0 else "\nH. Supp :\n0h00",
            ]
            if tot_gardes > 0:
                total_lines.append(f"\nGardes :\n{tot_gardes}")

            sc(ws.cell(row=base, column=COL_TOTAL),
               value="\n".join(total_lines),
               bg=TOTAL_BG, fg=DARK, size=9,
               wrap=True, halign='center', valign='center',
               border=b_outer(BLUE))
            ws.cell(row=base + 7, column=COL_TOTAL).fill   = PatternFill('solid', start_color=GREY_MID)
            ws.cell(row=base + 7, column=COL_TOTAL).border = b_bottom()

        ws.freeze_panes = f'{get_column_letter(COL_DAYS)}{HEADER_ROW + 1}'

        filename = f"resume_pointages_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    except Exception as e:
        return HttpResponse(f"Erreur lors de la génération du fichier : {str(e)}", status=500)

    return HttpResponse("Erreur inattendue : aucun retour explicite.", status=500)


# ---------------------------
# API VIEWSETS
# ---------------------------

from rest_framework import viewsets
from .serializers import (
    EmployeSerializer, SiteSerializer,
    PointageSerializer, PointageDetailSerializer
)


class EmployeViewSet(viewsets.ModelViewSet):
    queryset           = Employe.objects.filter(actif=True)
    serializer_class   = EmployeSerializer
    permission_classes = [IsAuthenticated]


class SiteViewSet(viewsets.ModelViewSet):
    queryset           = Site.objects.all()
    serializer_class   = SiteSerializer
    permission_classes = [IsAuthenticated]


class PointageViewSet(viewsets.ModelViewSet):
    queryset           = Pointage.objects.all()

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [IsAdminUser()]
        return [IsAuthenticated()]

    def get_serializer_class(self):
        return PointageDetailSerializer if self.action == 'retrieve' else PointageSerializer

    def get_queryset(self):
        queryset = Pointage.objects.all().select_related('employe', 'site')
        params   = self.request.query_params
        if params.get('date_debut'):   queryset = queryset.filter(date_pointage__gte=params['date_debut'])
        if params.get('date_fin'):     queryset = queryset.filter(date_pointage__lte=params['date_fin'])
        if params.get('employe_id'):   queryset = queryset.filter(employe_id=params['employe_id'])
        if params.get('site_id'):      queryset = queryset.filter(site_id=params['site_id'])
        if params.get('periode'):      queryset = queryset.filter(periode=params['periode'])
        if params.get('type_journee'): queryset = queryset.filter(type_journee=params['type_journee'])
        return queryset

    @api_view(['GET'])
    def statistiques(self, request):
        from datetime import timedelta
        today          = timezone.localtime(timezone.now()).date()
        total_employes = Employe.objects.filter(actif=True).count()
        data = {
            'total_employes':      total_employes,
            'presents_aujourdhui': Pointage.objects.filter(date_pointage=today).values('employe').distinct().count(),
            'absents_aujourdhui':  total_employes - Pointage.objects.filter(date_pointage=today).values('employe').distinct().count(),
            'retards_aujourdhui':  Pointage.objects.filter(date_pointage=today, periode__in=['matin', 'apres_midi'], retard__gt=timedelta(0)).count(),
            'gardes_en_cours':     Pointage.objects.filter(date_pointage=today, periode='nuit', type_journee='garde', heure_depart__isnull=True).count(),
            'date': today,
        }
        return Response(data)


class AnomaliePointageViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        return AnomaliePointageDetailSerializer if self.action == 'retrieve' else AnomaliePointageSerializer

    def get_queryset(self):
        queryset = AnomaliePointage.objects.select_related(
            'employe', 'site', 'traitement', 'traitement__administrateur', 'cloturee_par'
        )
        params = self.request.query_params
        if params.get('type'):
            queryset = queryset.filter(type=params['type'])
        if params.get('statut'):
            queryset = queryset.filter(statut=params['statut'])
        if params.get('employe_id'):
            queryset = queryset.filter(employe_id=params['employe_id'])
        if params.get('site_id'):
            queryset = queryset.filter(site_id=params['site_id'])
        return queryset

    @action(detail=True, methods=['post'])
    def traiter(self, request, pk=None):
        if not request.user.is_staff:
            return Response(
                {'success': False, 'error': 'Seul un administrateur ou RH peut traiter une anomalie.'},
                status=drf_status.HTTP_403_FORBIDDEN
            )
        
        anomalie = self.get_object()
        try:
            marquer_traitee(
                anomalie,
                administrateur=request.user,
                commentaire=request.data.get('commentaire', ''),
                corrections=request.data.get('corrections'),
            )
        except ValueError as e:
            return Response({'success': False, 'error': str(e)}, status=drf_status.HTTP_400_BAD_REQUEST)
        except PermissionError as e:
            return Response({'success': False, 'error': str(e)}, status=drf_status.HTTP_403_FORBIDDEN)
        
        anomalie.refresh_from_db()
        return Response(AnomaliePointageDetailSerializer(anomalie).data)

    @action(detail=True, methods=['post'])
    def cloturer(self, request, pk=None):
        if not request.user.is_staff:
            return Response(
                {'success': False, 'error': 'Seul un administrateur ou RH peut clôturer une anomalie.'},
                status=drf_status.HTTP_403_FORBIDDEN
            )
        
        anomalie = self.get_object()
        try:
            marquer_cloturee(anomalie, administrateur=request.user)
        except ValueError as e:
            return Response({'success': False, 'error': str(e)}, status=drf_status.HTTP_400_BAD_REQUEST)
        except PermissionError as e:
            return Response({'success': False, 'error': str(e)}, status=drf_status.HTTP_403_FORBIDDEN)
        
        anomalie.refresh_from_db()
        return Response(AnomaliePointageDetailSerializer(anomalie).data)


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework import status as drf_status

class ScanAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        raw_qr  = request.data.get('qr_data', '').strip()
        site_id = request.data.get('site_id')
        mode    = request.data.get('mode', 'auto')

        if not raw_qr or not site_id:
            return Response(
                {'success': False, 'error': 'qr_data et site_id sont requis'},
                status=drf_status.HTTP_400_BAD_REQUEST
            )

        parsed = parse_qr_data(raw_qr)
        if not parsed:
            return Response(
                {'success': False, 'error': 'Format QR invalide'},
                status=drf_status.HTTP_400_BAD_REQUEST
            )

        result = process_scan(
            matricule=parsed['matricule'],
            qr_token=parsed['token'],
            site_id=int(site_id),
            mode=mode
        )

        if result['status'] == 'success':
            return Response({'success': True, **result}, status=drf_status.HTTP_201_CREATED)
        elif result['status'] == 'warning':
            return Response({'success': False, **result}, status=drf_status.HTTP_200_OK)
        else:
            return Response({'success': False, **result}, status=drf_status.HTTP_400_BAD_REQUEST)


# ---------------------------
# FONCTIONS API SUPPLÉMENTAIRES
# ---------------------------

from rest_framework.decorators import api_view, permission_classes as pc


@api_view(['POST'])
@pc([IsAuthenticated])
def scan_api_view(request):
    raw_qr  = request.data.get('qr_data', '').strip()
    site_id = request.data.get('site_id')
    mode    = request.data.get('mode', 'auto')

    if not raw_qr or not site_id:
        return Response(
            {'success': False, 'error': 'qr_data et site_id requis'},
            status=drf_status.HTTP_400_BAD_REQUEST
        )

    parsed = parse_qr_data(raw_qr)
    if not parsed:
        return Response({'success': False, 'error': 'Format QR invalide'}, status=400)

    result = process_scan(
        matricule=parsed['matricule'],
        qr_token=parsed['token'],
        site_id=int(site_id),
        mode=mode
    )

    http = 201 if result['status'] == 'success' else (
        200 if result['status'] == 'warning' else 400
    )
    return Response({'success': result['status'] == 'success', **result}, status=http)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_statut_journee(request, employe_id):
    try:
        employe       = Employe.objects.get(id=employe_id, actif=True)
        date_str      = request.GET.get('date')
        date_courante = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else timezone.localtime(timezone.now()).date()
        pointages     = Pointage.objects.filter(employe=employe, date_pointage=date_courante)

        statut = {
            'date': date_courante, 'employe_id': employe_id,
            'matin':      {'present': False, 'heure_arrivee': None, 'heure_depart': None, 'site': None},
            'apres_midi': {'present': False, 'heure_arrivee': None, 'heure_depart': None, 'site': None},
            'nuit':       {'present': False, 'heure_arrivee': None, 'heure_depart': None, 'site': None, 'type_journee': None},
            'heures_travaillees': str(timedelta(0)), 'heures_supplementaires': str(timedelta(0)),
        }

        def _add_heures(current_str, delta):
            parts   = current_str.split(':')
            current = timedelta(seconds=sum(int(x) * 60**i for i, x in enumerate(reversed(parts))))
            return str(current + delta)

        for p in pointages:
            k = p.periode if p.periode in ('matin', 'apres_midi', 'nuit') else None
            if k:
                statut[k].update({'present': True, 'heure_arrivee': p.heure_arrivee, 'heure_depart': p.heure_depart, 'site': p.site.nom if p.site else None})
                if k == 'nuit':
                    statut[k]['type_journee'] = p.type_journee
                if p.heures_travaillees:
                    statut['heures_travaillees'] = _add_heures(statut['heures_travaillees'], p.heures_travaillees)

        parts = statut['heures_travaillees'].split(':')
        total = timedelta(seconds=sum(int(x) * 60**i for i, x in enumerate(reversed(parts))))
        if total > timedelta(hours=8):
            statut['heures_supplementaires'] = str(total - timedelta(hours=8))

        return Response(statut)
    except Employe.DoesNotExist:
        return Response({'error': 'Employé non trouvé'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_prochain_scan(request, employe_id):
    try:
        employe = Employe.objects.get(id=employe_id, actif=True)
        today   = timezone.localtime(timezone.now()).date()

        if Pointage.objects.filter(employe=employe, date_pointage=today, periode='nuit', type_journee='garde', heure_depart__isnull=True).exists():
            return Response({'prochain_scan': 'fin_garde',   'type': 'garde',    'message': 'Fin de garde attendue'})
        if Pointage.objects.filter(employe=employe, date_pointage=today, periode='nuit', type_journee='garde', heure_arrivee__isnull=True).exists():
            return Response({'prochain_scan': 'debut_garde', 'type': 'garde',    'message': 'Début de garde attendu'})

        maintenant = timezone.localtime(timezone.now()).time()
        periode    = 'apres_midi' if maintenant >= time(12, 0) else 'matin'
        pointage   = Pointage.objects.filter(employe=employe, date_pointage=today, periode=periode).first()

        if not pointage or not pointage.heure_arrivee:
            return Response({'prochain_scan': f'entree_{periode}', 'type': 'pointage', 'periode': periode, 'message': f"Entrée {periode} attendue"})
        if not pointage.heure_depart:
            return Response({'prochain_scan': f'sortie_{periode}', 'type': 'pointage', 'periode': periode, 'message': f"Sortie {periode} attendue"})

        if periode == 'matin' and maintenant >= time(12, 0):
            pa = Pointage.objects.filter(employe=employe, date_pointage=today, periode='apres_midi').first()
            if not pa or not pa.heure_arrivee:
                return Response({'prochain_scan': 'entree_apres_midi', 'type': 'pointage', 'periode': 'apres_midi', 'message': "Entrée après-midi attendue"})
            if not pa.heure_depart:
                return Response({'prochain_scan': 'sortie_apres_midi', 'type': 'pointage', 'periode': 'apres_midi', 'message': "Sortie après-midi attendue"})

        return Response({'prochain_scan': None, 'type': 'complet', 'message': "Tous les pointages pour aujourd'hui sont terminés"})
    except Employe.DoesNotExist:
        return Response({'error': 'Employé non trouvé'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def employe_qr_data(request, matricule):
    try:
        employe = Employe.objects.get(matricule=matricule, actif=True)
        return Response({
            'matricule': employe.matricule, 'nom': employe.nom, 'prenom': employe.prenom,
            'poste': employe.poste.nom if employe.poste else "Non défini",
            'qr_token': str(employe.qr_code_token)
        })
    except Employe.DoesNotExist:
        return Response({'error': 'Employé non trouvé'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_dashboard_stats(request):
    today          = timezone.localtime(timezone.now()).date()
    total_employes = Employe.objects.filter(actif=True).count()
    jours_presents = [Pointage.objects.filter(date_pointage=today - timedelta(days=i)).values('employe').distinct().count() for i in range(6, -1, -1)]
    return JsonResponse({'success': True, 'data': {
        'total_employes':      total_employes,
        'presents_aujourdhui': Pointage.objects.filter(date_pointage=today).values('employe').distinct().count(),
        'absents_aujourdhui':  total_employes - Pointage.objects.filter(date_pointage=today).values('employe').distinct().count(),
        'retards_aujourdhui':  Pointage.objects.filter(date_pointage=today, periode__in=['matin', 'apres_midi'], retard__gt=timedelta(0)).count(),
        'gardes_en_cours':     Pointage.objects.filter(date_pointage=today, periode='nuit', type_journee='garde', heure_depart__isnull=True).count(),
        'anomalies_ouvertes':  compter_anomalies_ouvertes(),
        'weekly_data':         jours_presents,
        'timestamp':           timezone.localtime(timezone.now()).isoformat(),
    }})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_charts_data(request):
    today          = timezone.localtime(timezone.now()).date()
    total_employes = Employe.objects.filter(actif=True).count()

    jours_labels, jours_presents, jours_retards = [], [], []
    for i in range(6, -1, -1):
        jour = today - timedelta(days=i)
        jours_labels.append(jour.strftime('%a'))
        jours_presents.append(Pointage.objects.filter(date_pointage=jour).values('employe').distinct().count())
        jours_retards.append(Pointage.objects.filter(date_pointage=jour, periode__in=['matin', 'apres_midi'], retard__gt=timedelta(0)).count())

    semaines_labels, semaines_taux = [], []
    for i in range(3, -1, -1):
        start = today - timedelta(days=today.weekday() + 7*i)
        end   = start + timedelta(days=6)
        semaines_labels.append(f"Sem. {start.strftime('%d/%m')}")
        emp   = Pointage.objects.filter(date_pointage__gte=start, date_pointage__lte=end).values('employe').distinct().count()
        semaines_taux.append(round((emp / total_employes * 100) if total_employes else 0, 1))

    presents = Pointage.objects.filter(date_pointage=today).values('employe').distinct().count()
    retards  = Pointage.objects.filter(date_pointage=today, periode__in=['matin', 'apres_midi'], retard__gt=timedelta(0)).count()
    gardes   = Pointage.objects.filter(date_pointage=today, periode='nuit', type_journee='garde', heure_depart__isnull=True).count()

    return JsonResponse({'success': True,
        'daily':     {'presents': presents, 'absents': total_employes - presents, 'retards': retards, 'gardes': gardes, 'labels': ['Présents', 'Absents', 'Retards', 'Gardes']},
        'weekly':    {'labels': jours_labels, 'presents': jours_presents, 'retards': jours_retards},
        'evolution': {'labels': semaines_labels, 'taux_presence': semaines_taux},
    })
# pointage/views.py

from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import DemandeModification, AnomaliePointage


@login_required
def admin_badge_counts_api(request):
    """API pour les compteurs de badges admin"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    return JsonResponse({
        'demandes_attente': DemandeModification.objects.filter(statut='en_attente').count(),
        'anomalies_ouvertes': AnomaliePointage.objects.filter(
            statut=AnomaliePointage.STATUT_OUVERTE
        ).count(),
    })