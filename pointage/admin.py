# pointage/admin.py - Version complète avec tous les filtres UI

from django.contrib import admin
from django.contrib.admin import SimpleListFilter
from django.utils.safestring import mark_safe
from django.utils.html import format_html
from django.urls import path, reverse
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib import messages
from django.contrib.auth.admin import UserAdmin
from django.utils import timezone
from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import json
from .models import (
    Employe, Site, Pointage, Scan, Poste,
    CustomUser, DemandeModification,
    AnomaliePointage, AnomalieTraitement,
)
from .anomalies import marquer_traitee, marquer_cloturee
from .forms import PointageForm
import uuid
from datetime import timedelta, datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# FILTRES EXACTEMENT COMME L'INTERFACE UTILISATEUR
# ============================================================

class DateInputFilterMixin:
    """Filtre admin affichant un champ <input type="date"> au lieu d'une liste
    de choix prédéfinis, pour permettre une date arbitraire (comme dans l'UI).

    IMPORTANT : SimpleListFilter n'affiche la sidebar / n'appelle queryset()
    que si has_output() renvoie True. Par défaut has_output() renvoie
    `len(self.lookup_choices) > 0`, donc avec lookups() vide (nécessaire ici
    puisqu'on ne veut pas de liste de choix), Django ignore silencieusement
    le filtre. On force donc has_output() à True.
    """
    template = 'admin/pointage/date_input_filter.html'

    def lookups(self, request, model_admin):
        return ()

    def has_output(self):
        return True

    def choices(self, changelist):
        hidden_fields = [
            (key, value) for key, value in changelist.params.items()
            if key != self.parameter_name
        ]
        yield {
            'value': self.value() or '',
            'parameter_name': self.parameter_name,
            'hidden_fields': hidden_fields,
            'reset_query_string': changelist.get_query_string(remove=[self.parameter_name]),
        }


class DateDebutFilter(DateInputFilterMixin, SimpleListFilter):
    """Date début - retourne les pointages à partir de cette date (incluse)."""
    title = 'Date début'
    parameter_name = 'date_debut'

    def queryset(self, request, queryset):
        value = self.value()
        if not value:
            return queryset
        try:
            date_debut = datetime.strptime(value, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return queryset
        # date_pointage est un DateField (pas DateTimeField) : gte est déjà
        # inclusif sur la journée entière, aucun souci de fuseau horaire ici.
        return queryset.filter(date_pointage__gte=date_debut)


class DateFinFilter(DateInputFilterMixin, SimpleListFilter):
    """Date fin - retourne les pointages jusqu'à cette date (incluse)."""
    title = 'Date fin'
    parameter_name = 'date_fin'

    def queryset(self, request, queryset):
        value = self.value()
        if not value:
            return queryset
        try:
            date_fin = datetime.strptime(value, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return queryset
        return queryset.filter(date_pointage__lte=date_fin)


class EmployeFilter(SimpleListFilter):
    """Employé - comme dans l'UI"""
    title = 'Employé'
    parameter_name = 'employe'

    def lookups(self, request, model_admin):
        employes = Employe.objects.filter(actif=True).order_by('nom', 'prenom')
        return [(str(e.id), f"{e.prenom} {e.nom} ({e.matricule})") for e in employes]

    def queryset(self, request, queryset):
        if self.value():
            try:
                return queryset.filter(employe_id=int(self.value()))
            except (ValueError, TypeError):
                return queryset
        return queryset


class SiteFilter(SimpleListFilter):
    """Site - comme dans l'UI"""
    title = 'Site'
    parameter_name = 'site'

    def lookups(self, request, model_admin):
        sites = Site.objects.all().order_by('nom')
        return [(str(s.id), s.nom) for s in sites]

    def queryset(self, request, queryset):
        if self.value():
            try:
                return queryset.filter(site_id=int(self.value()))
            except (ValueError, TypeError):
                return queryset
        return queryset


class PeriodeTypeFilter(SimpleListFilter):
    """Type de période - comme dans l'UI (Jour/Nuit)"""
    title = 'Type de période'
    parameter_name = 'periode_type'

    def lookups(self, request, model_admin):
        return [
            ('jour', 'Jour'),
            ('nuit', 'Nuit (Gardes)'),
        ]

    def queryset(self, request, queryset):
        if self.value() == 'jour':
            return queryset.filter(type_journee='normal')
        if self.value() == 'nuit':
            return queryset.filter(type_journee='garde')
        return queryset


# ============================================================
# CUSTOM USER
# ============================================================

@admin.register(CustomUser)
class CustomUserAdmin(UserAdmin):
    list_display = ('username', 'email', 'first_name', 'last_name', 'role', 'is_active', 'is_staff')
    list_filter = ('role', 'is_active', 'is_staff')
    search_fields = ('username', 'email', 'first_name', 'last_name')
    ordering = ('username',)

    fieldsets = UserAdmin.fieldsets + (
        ('Rôle & Permissions', {'fields': ('role',)}),
    )
    add_fieldsets = UserAdmin.add_fieldsets + (
        ('Rôle & Permissions', {'fields': ('role',)}),
    )

    def save_model(self, request, obj, form, change):
        if obj.role == 'admin':
            obj.is_staff = True
        else:
            obj.is_staff = False
            obj.is_superuser = False
        super().save_model(request, obj, form, change)


# ============================================================
# POSTE
# ============================================================

@admin.register(Poste)
class PosteAdmin(admin.ModelAdmin):
    list_display = ('nom', 'description', 'couleur_display')
    search_fields = ('nom', 'description')
    ordering = ('nom',)

    def couleur_display(self, obj):
        return format_html(
            '<span style="display:inline-block;width:20px;height:20px;'
            'background-color:{};border:1px solid rgba(255,255,255,.2);'
            'border-radius:4px;vertical-align:middle;margin-right:6px"></span>{}',
            obj.couleur, obj.couleur
        )
    couleur_display.short_description = 'Couleur'


# ============================================================
# SITE
# ============================================================

@admin.register(Site)
class SiteAdmin(admin.ModelAdmin):
    list_display = ('nom', 'adresse', 'heure_ouverture_matin', 'heure_fermeture_matin')
    search_fields = ('nom', 'adresse')


# ============================================================
# EMPLOYÉ
# ============================================================

@admin.register(Employe)
class EmployeAdmin(admin.ModelAdmin):
    list_display = ('matricule', 'nom', 'prenom', 'get_poste', 'actif', 'qr_code_preview', 'date_creation')
    list_filter = ('poste', 'actif', 'date_creation')
    search_fields = ('nom', 'prenom', 'matricule', 'poste__nom')
    readonly_fields = ('qr_code_token', 'date_creation', 'qr_code_display', 'info_qr_code')
    ordering = ('matricule',)
    actions = ['regenerer_qr_codes', 'activer_employes', 'desactiver_employes']

    fieldsets = (
        ('Informations personnelles', {
            'fields': ('nom', 'prenom', 'matricule', 'poste', 'actif')
        }),
        ('QR Code', {
            'fields': ('info_qr_code', 'qr_code_display', 'qr_code_token'),
            'classes': ('wide',),
        }),
        ('Dates', {
            'fields': ('date_creation',),
            'classes': ('collapse',),
        }),
    )

    def qr_code_preview(self, obj):
        if obj.qr_code:
            return format_html(
                '<a href="{}" target="_blank">'
                '<img src="{}" width="40" height="40" '
                'style="border:1px solid rgba(255,255,255,.15);border-radius:4px;">'
                '</a>',
                obj.qr_code.url, obj.qr_code.url
            )
        return mark_safe('<span style="color:#f87171;font-size:12px;">Non généré</span>')
    qr_code_preview.short_description = 'QR Code'

    def qr_code_display(self, obj):
        if obj.qr_code:
            return format_html(
                '<div style="text-align:center;margin:20px 0;">'
                '<div style="margin-bottom:12px;font-weight:600;color:#e8eaf0;">QR Code</div>'
                '<img src="{}" width="220" height="220" '
                'style="border:3px solid #4f8ef7;border-radius:10px;padding:10px;background:white;">'
                '</div>',
                obj.qr_code.url
            )
        return mark_safe('<div style="color:#f87171;padding:14px;text-align:center;">Non généré</div>')
    qr_code_display.short_description = 'Visualisation QR Code'

    def info_qr_code(self, obj):
        if obj.qr_code:
            return format_html(
                '<div style="background:rgba(79,142,247,.08);padding:16px;border-radius:8px;'
                'border:1px solid rgba(79,142,247,.2);margin-bottom:14px;">'
                '<h4 style="margin-top:0;color:#4f8ef7;font-size:13px;font-weight:600;">QR Code</h4>'
                '<code style="background:rgba(255,255,255,.07);padding:6px 10px;border-radius:5px;'
                'font-size:12px;color:#e8eaf0;display:inline-block;margin-bottom:14px;">'
                'EMPLOYE:{}:{}</code>'
                '<div style="display:flex;gap:8px;flex-wrap:wrap;">'
                '<a href="{}" download class="button" style="text-decoration:none;">Télécharger</a>'
                '<a href="{}" target="_blank" class="button" style="text-decoration:none;">Ouvrir</a>'
                '<button type="submit" name="_generate_qr" value="1" class="button" '
                'style="background:#28a745;border-color:#28a745;">Régénérer</button>'
                '</div></div>',
                obj.matricule, obj.qr_code_token, obj.qr_code.url, obj.qr_code.url
            )
        return mark_safe(
            '<div style="background:rgba(251,191,36,.08);padding:14px;border-radius:8px;'
            'border:1px solid rgba(251,191,36,.2);">'
            '<button type="submit" name="_generate_qr" value="1" class="button" '
            'style="background:#4f8ef7;border-color:#4f8ef7;color:white;">Générer QR Code</button>'
            '</div>'
        )
    info_qr_code.short_description = 'Actions QR Code'

    def get_poste(self, obj):
        return obj.poste.nom if obj.poste else "Non défini"
    get_poste.short_description = 'Poste'

    def response_change(self, request, obj):
        if "_generate_qr" in request.POST:
            obj.qr_code_token = uuid.uuid4()
            obj.generer_qr_code()
            obj.save()
            self.message_user(request, "✅ QR code régénéré avec succès !", messages.SUCCESS)
            return HttpResponseRedirect(".")
        return super().response_change(request, obj)

    def regenerer_qr_codes(self, request, queryset):
        count = 0
        for employe in queryset:
            employe.qr_code_token = uuid.uuid4()
            employe.generer_qr_code()
            employe.save()
            count += 1
        self.message_user(request, f"✅ {count} QR code(s) régénéré(s).", messages.SUCCESS)
    regenerer_qr_codes.short_description = "🔄 Régénérer les QR codes"

    def activer_employes(self, request, queryset):
        updated = queryset.update(actif=True)
        self.message_user(request, f"✅ {updated} employé(s) activé(s).", messages.SUCCESS)
    activer_employes.short_description = "✅ Activer"

    def desactiver_employes(self, request, queryset):
        updated = queryset.update(actif=False)
        self.message_user(request, f"⛔ {updated} employé(s) désactivé(s).", messages.SUCCESS)
    desactiver_employes.short_description = "⛔ Désactiver"

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('poste')


# ============================================================
# POINTAGE - AVEC FILTRES EXACTEMENT COMME L'UI
# ============================================================

@admin.register(Pointage)
class PointageAdmin(admin.ModelAdmin):
    change_list_template = "admin/pointage/pointage_changelist.html"
    
    list_display = [
        'employe',
        'date_pointage',
        'periode',
        'type_journee',
        'heure_arrivee',
        'heure_depart',
        'site',
        'statut',
        'get_retard_display',
        'get_heures_display',
    ]
    
    # ============================================================
    # FILTRES EXACTEMENT COMME L'INTERFACE UTILISATEUR
    # ============================================================
    list_filter = [
        DateDebutFilter,      # Date début
        DateFinFilter,        # Date fin
        EmployeFilter,        # Employé
        SiteFilter,           # Site
        PeriodeTypeFilter,    # Type de période (Jour/Nuit)
    ]
    
    search_fields = [
        'employe__nom',
        'employe__prenom',
        'employe__matricule',
        'site__nom',
    ]
    
    readonly_fields = ('retard', 'heures_travaillees', 'date_creation', 'date_modification')
    date_hierarchy = 'date_pointage'
    
    # ============================================================
    # CHAMPS PERSONNALISÉS
    # ============================================================
    
    def get_retard_display(self, obj):
        if obj.retard and obj.retard.total_seconds() > 0:
            minutes = obj.get_retard_minutes()
            if minutes >= 30:
                return format_html(
                    '<span style="background:#FEE2E2;color:#B91C1C;padding:2px 10px;border-radius:9999px;font-weight:600;font-size:10px;">⚠️ {} min</span>',
                    minutes
                )
            return f"{minutes} min"
        return "—"
    get_retard_display.short_description = "Retard"
    
    def get_heures_display(self, obj):
        if obj.heures_travaillees and obj.heures_travaillees.total_seconds() > 0:
            total_seconds = obj.heures_travaillees.total_seconds()
            hours = int(total_seconds // 3600)
            minutes = int((total_seconds % 3600) // 60)
            return f"{hours}h{minutes:02d}"
        return "—"
    get_heures_display.short_description = "Heures travaillées"
    
    # ============================================================
    # ACTIONS EN MASSE
    # ============================================================
    
    actions = ['marquer_present', 'marquer_retard', 'marquer_absent', 'supprimer_selection']
    
    def marquer_present(self, request, queryset):
        queryset.update(statut='present')
        self.message_user(request, f"✅ {queryset.count()} pointage(s) marqué(s) comme présent.")
    marquer_present.short_description = "✅ Marquer comme présent"
    
    def marquer_retard(self, request, queryset):
        queryset.update(statut='retard')
        self.message_user(request, f"⚠️ {queryset.count()} pointage(s) marqué(s) comme retard.")
    marquer_retard.short_description = "⚠️ Marquer comme retard"
    
    def marquer_absent(self, request, queryset):
        queryset.update(statut='absent')
        self.message_user(request, f"❌ {queryset.count()} pointage(s) marqué(s) comme absent.")
    marquer_absent.short_description = "❌ Marquer comme absent"
    
    def supprimer_selection(self, request, queryset):
        count = queryset.count()
        if count > 0:
            queryset.delete()
            self.message_user(request, f"🗑️ {count} pointage(s) supprimé(s).")
    supprimer_selection.short_description = "🗑️ Supprimer"
    
    # ============================================================
    # GET QUERYSET
    # ============================================================
    
    def get_queryset(self, request):
        return Pointage.objects.select_related('employe', 'site')
    
    # ============================================================
    # CHANGELIST VIEW
    # ============================================================
    
    def changelist_view(self, request, extra_context=None):
        # 'cards_page' est un paramètre de pagination propre à la vue
        # "Cartes" ci-dessous. Django's ChangeList interprète tout paramètre
        # GET non reconnu (hors recherche/tri/pagination native/list_filter)
        # comme un lookup de champ à appliquer au queryset -> il faut donc
        # le retirer AVANT de construire le ChangeList, sous peine de
        # IncorrectLookupParameters ("Cannot resolve keyword 'cards_page'").
        cards_page_value = request.GET.get('cards_page')
        if 'cards_page' in request.GET:
            cleaned_get = request.GET.copy()
            cleaned_get.pop('cards_page')
            request.GET = cleaned_get

        cl = self.get_changelist_instance(request)
        queryset = cl.get_queryset(request)
        
        if 'export_excel' in request.GET:
            return self.export_excel(request, queryset)
        
        total = queryset.count()
        presents = queryset.filter(statut='present').count()
        retards = queryset.filter(statut='retard').count()
        absents = queryset.filter(statut='absent').count()
        
        extra_context = extra_context or {}
        extra_context.update({
            'total': total,
            'presents_count': presents,
            'retards_count': retards,
            'absents_count': absents,
            'jours_cards': self._build_jours_cards(queryset, cards_page_value),
        })
        
        return super().changelist_view(request, extra_context=extra_context)

    def _build_jours_cards(self, queryset, cards_page_value):
        """Regroupe le queryset (déjà filtré par la sidebar admin) par
        (employé, jour), comme l'affichage 'Historique' côté utilisateur —
        avec quelques améliorations : statut 'partiel' réellement calculé
        (pas seulement présent/absent), et prefetch pour éviter le N+1.
        """
        pointages = list(
            queryset.select_related('employe', 'employe__poste', 'site')
            .prefetch_related('scans', 'scans__site')
            .order_by('-date_pointage', 'employe__nom')[:500]
        )

        jours_dict = {}
        for pointage in pointages:
            key = (pointage.employe_id, pointage.date_pointage)
            if key not in jours_dict:
                jours_dict[key] = {
                    'date': pointage.date_pointage, 'employe': pointage.employe, 'site': pointage.site,
                    'matin': None, 'apres_midi': None, 'nuit': None,
                }
            jours_dict[key][pointage.periode] = pointage

        jours_list = []
        for jour in jours_dict.values():
            heures_total = timedelta()
            retard_total = timedelta()
            segments_attendus = 0
            segments_complets = 0
            for cle in ('matin', 'apres_midi'):
                p = jour[cle]
                if p:
                    segments_attendus += 1
                    heures_total += p.heures_travaillees or timedelta()
                    retard_total += p.retard or timedelta()
                    if p.heure_arrivee and p.heure_depart:
                        segments_complets += 1
            if jour['nuit']:
                heures_total += jour['nuit'].heures_travaillees or timedelta()

            if jour['nuit']:
                statut_global = 'present' if (jour['nuit'].heure_arrivee and jour['nuit'].heure_depart) else 'partiel'
            elif segments_attendus == 0:
                statut_global = 'absent'
            elif segments_complets == segments_attendus:
                statut_global = 'present'
            else:
                statut_global = 'partiel'

            jour['heures_total']  = heures_total
            jour['retard_total']  = retard_total
            jour['heures_sup']    = max(heures_total - timedelta(hours=8), timedelta())
            jour['statut_global'] = statut_global
            jour['is_garde']      = bool(jour['nuit'] and jour['nuit'].type_journee == 'garde')
            jour['multisite']     = bool(
                jour['matin'] and jour['apres_midi'] and jour['matin'].site_id != jour['apres_midi'].site_id
            )
            # ID du Pointage à utiliser pour le lien "Voir dans l'admin"
            ref = jour['nuit'] or jour['matin'] or jour['apres_midi']
            jour['admin_pointage_id'] = ref.id if ref else None
            jours_list.append(jour)

        jours_list.sort(key=lambda j: (j['date'], j['employe'].nom), reverse=True)

        paginator = Paginator(jours_list, 12)
        try:
            page_number = int(cards_page_value or 1)
        except (TypeError, ValueError):
            page_number = 1
        try:
            jours_page = paginator.page(page_number)
        except (EmptyPage, PageNotAnInteger):
            jours_page = paginator.page(1)
        return jours_page
    
    # ============================================================
    # EXPORT EXCEL
    # ============================================================
    
    def export_excel(self, request, queryset):
        """Export Excel au même format que l'interface utilisateur"""
        if not queryset.exists():
            messages.warning(request, "Aucun pointage à exporter.")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/admin/pointage/pointage/'))
        
        dates = queryset.values_list('date_pointage', flat=True).distinct().order_by('date_pointage')
        if not dates:
            messages.warning(request, "Aucune date trouvée.")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/admin/pointage/pointage/'))
        
        date_debut = dates[0]
        date_fin = dates.last()
        
        emp_data = defaultdict(lambda: defaultdict(lambda: {'matin': None, 'apres_midi': None, 'nuit': None}))
        emp_info = {}
        
        for p in queryset:
            emp_info[p.employe.id] = (p.employe.id, p.employe.get_nom_complet(), p.employe.matricule)
            emp_data[p.employe.id][p.date_pointage][p.periode] = p
        
        def work_days(d1, d2):
            days, d = [], d1
            while d <= d2:
                days.append(d)
                d += timedelta(days=1)
            return days
        
        days = work_days(date_debut, date_fin)
        JOURS_FR = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
        
        def fmt_duree(pointage):
            if not pointage or not pointage.heures_travaillees:
                return '—'
            return pointage.get_duree_formatee()
        
        def fmt_time(t):
            return t.strftime('%H:%M') if t else '—'
        
        BLUE = '1E3A5F'
        BLUE_LIGHT = 'D6E4F0'
        ORANGE_BG = 'FEF3C7'
        ORANGE_FG = 'D97706'
        GREEN_BG = 'DCFCE7'
        GREEN_FG = '15803D'
        RED_BG = 'FEE2E2'
        RED_FG = 'B91C1C'
        PURPLE_BG = 'EDE9FE'
        PURPLE_FG = '7C3AED'
        PURPLE_MED = '4C1D95'
        NIGHT_BG = '1E1B4B'
        NIGHT_MID = '2D1F4E'
        NIGHT_FG = 'A5B4FC'
        DARK = '1A1A1A'
        GREY_LIGHT = 'F5F5F7'
        GREY_MID = 'E5E5E5'
        WHITE = 'FFFFFF'
        TOTAL_BG = 'EEF2FF'
        
        def sd(color=GREY_MID, style='thin'):
            return Side(style=style, color=color)
        
        def b_all(color=GREY_MID):
            s = sd(color)
            return Border(left=s, right=s, top=s, bottom=s)
        
        def b_outer(color=BLUE):
            s = Side(style='medium', color=color)
            return Border(left=s, right=s, top=s, bottom=s)
        
        def b_bottom(color=BLUE):
            return Border(left=sd(), right=sd(), top=sd(),
                          bottom=Side(style='medium', color=color))
        
        def sc(c, value='', bg=WHITE, fg=DARK, bold=False, size=9,
               halign='center', valign='center', wrap=False, border=None, italic=False):
            c.value = value
            c.font = Font(name='Arial', bold=bold, color=fg, size=size, italic=italic)
            c.fill = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
            if border:
                c.border = border
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Résumé Pointages"
        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        
        COL_EMP = 1
        COL_DAYS = 2
        COL_TOTAL = COL_DAYS + len(days)
        
        ws.column_dimensions[get_column_letter(COL_EMP)].width = 15
        for i in range(len(days)):
            ws.column_dimensions[get_column_letter(COL_DAYS + i)].width = 14
        ws.column_dimensions[get_column_letter(COL_TOTAL)].width = 18
        
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COL_TOTAL)
        sc(ws['A1'],
           value=f"RÉSUMÉ DES POINTAGES  ·  Du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}",
           bg=BLUE, fg=WHITE, bold=True, size=13)
        ws.row_dimensions[1].height = 36
        
        HEADER_ROW = 2
        sc(ws.cell(row=HEADER_ROW, column=COL_EMP), value='Employé',
           bg=BLUE, fg=WHITE, bold=True, size=9, border=b_all(BLUE))
        for i, d in enumerate(days):
            label = f"{JOURS_FR[d.weekday()]}\n{d.strftime('%d/%m/%Y')}"
            sc(ws.cell(row=HEADER_ROW, column=COL_DAYS + i), value=label,
               bg=BLUE, fg=WHITE, bold=True, size=8, wrap=True, border=b_all(BLUE))
        sc(ws.cell(row=HEADER_ROW, column=COL_TOTAL), value='TOTAL',
           bg=DARK, fg=WHITE, bold=True, size=10, border=b_all(DARK))
        ws.row_dimensions[HEADER_ROW].height = 28
        
        ROWS_PER_EMP = 8
        SUB_H = [8, 14, 18, 14, 18, 16, 16, 6]
        START_ROW = 3
        
        for emp_idx, (emp_id, day_map) in enumerate(emp_data.items()):
            base = START_ROW + emp_idx * ROWS_PER_EMP
            _, nom_complet, matricule = emp_info[emp_id]
            
            for i, h in enumerate(SUB_H):
                ws.row_dimensions[base + i].height = h
            
            ws.merge_cells(start_row=base, start_column=COL_EMP,
                           end_row=base + 6, end_column=COL_EMP)
            sc(ws.cell(row=base, column=COL_EMP),
               value=f"{nom_complet}\n{matricule}",
               bg=GREY_LIGHT, fg=DARK, bold=True, size=9,
               wrap=True, halign='center', valign='center', border=b_outer())
            ws.cell(row=base + 7, column=COL_EMP).fill = PatternFill('solid', start_color=GREY_MID)
            ws.cell(row=base + 7, column=COL_EMP).border = b_bottom()
            
            tot_retard = timedelta()
            tot_trav = timedelta()
            tot_sup = timedelta()
            tot_gardes = 0
            
            for i, d in enumerate(days):
                col = COL_DAYS + i
                pt_map = day_map.get(d, {})
                matin = pt_map.get('matin')
                apm = pt_map.get('apres_midi')
                nuit = pt_map.get('nuit')
                bg_day = WHITE if i % 2 == 0 else 'FAFAFA'
                
                if nuit and nuit.type_journee == 'garde':
                    tot_gardes += 1
                    h_garde = nuit.heures_travaillees or timedelta()
                    tot_trav += h_garde
                    terminee = bool(nuit.heure_depart)
                    
                    sc(ws.cell(row=base, column=col), bg=NIGHT_MID,
                       border=Border(top=Side(style='medium', color=PURPLE_FG),
                                     left=sd(), right=sd()))
                    ws.merge_cells(start_row=base + 1, start_column=col,
                                   end_row=base + 2, end_column=col)
                    sc(ws.cell(row=base + 1, column=col),
                       value='🌙 Garde de nuit',
                       bg=PURPLE_BG, fg=PURPLE_FG, bold=True, size=9,
                       wrap=True, halign='center', valign='center',
                       border=b_all(PURPLE_FG))
                    sc(ws.cell(row=base + 3, column=col),
                       value='Début  →  Fin',
                       bg=NIGHT_MID, fg=NIGHT_FG, bold=True, size=8,
                       border=b_all(PURPLE_MED))
                    arr_g = fmt_time(nuit.heure_arrivee)
                    dep_g = fmt_time(nuit.heure_depart)
                    sc(ws.cell(row=base + 4, column=col),
                       value=f"{arr_g}  →  {dep_g}",
                       bg=NIGHT_BG, fg=WHITE, bold=True, size=9,
                       border=b_all(PURPLE_MED))
                    sc(ws.cell(row=base + 5, column=col),
                       value=f"Durée : {fmt_duree(nuit)}",
                       bg=PURPLE_BG, fg=PURPLE_FG, bold=True, size=8,
                       border=b_all(PURPLE_FG))
                    if terminee:
                        sc(ws.cell(row=base + 6, column=col),
                           value='✓ Terminée', bg=GREEN_BG, fg=GREEN_FG,
                           bold=True, size=8, border=b_all())
                    else:
                        sc(ws.cell(row=base + 6, column=col),
                           value='⏳ En cours', bg=ORANGE_BG, fg=ORANGE_FG,
                           bold=True, size=8, border=b_all())
                    sc(ws.cell(row=base + 7, column=col), bg=GREY_LIGHT,
                       border=Border(bottom=Side(style='medium', color=PURPLE_FG),
                                     left=sd(), right=sd()))
                else:
                    has_data = matin or apm
                    h_trav = timedelta()
                    h_ret = timedelta()
                    if matin:
                        h_trav += matin.heures_travaillees or timedelta()
                        h_ret += matin.retard or timedelta()
                    if apm:
                        h_trav += apm.heures_travaillees or timedelta()
                        h_ret += apm.retard or timedelta()
                    h_sup = max(timedelta(), h_trav - timedelta(hours=8))
                    
                    tot_trav += h_trav
                    tot_retard += h_ret
                    tot_sup += h_sup
                    
                    sc(ws.cell(row=base, column=col),
                       bg=bg_day if has_data else GREY_LIGHT,
                       border=Border(top=Side(style='medium', color=BLUE),
                                     left=sd(), right=sd()))
                    sc(ws.cell(row=base + 1, column=col),
                       value='Matin', bg=ORANGE_BG, fg=ORANGE_FG,
                       bold=True, size=8, border=b_all())
                    arr_m = fmt_time(matin.heure_arrivee if matin else None)
                    dep_m = fmt_time(matin.heure_depart if matin else None)
                    sc(ws.cell(row=base + 2, column=col),
                       value=f"{arr_m}  →  {dep_m}" if has_data else '—',
                       bg=bg_day, fg=DARK, bold=True, size=9, border=b_all())
                    sc(ws.cell(row=base + 3, column=col),
                       value='Après-midi', bg=BLUE_LIGHT, fg=BLUE,
                       bold=True, size=8, border=b_all())
                    arr_s = fmt_time(apm.heure_arrivee if apm else None)
                    dep_s = fmt_time(apm.heure_depart if apm else None)
                    sc(ws.cell(row=base + 4, column=col),
                       value=f"{arr_s}  →  {dep_s}" if has_data else '—',
                       bg=bg_day, fg=DARK, bold=True, size=9, border=b_all())
                    sc(ws.cell(row=base + 5, column=col),
                       value=f"Retard : {int(h_ret.total_seconds() // 60)}min" if h_ret.total_seconds() > 0 else '—',
                       bg=RED_BG if h_ret.total_seconds() > 0 else bg_day,
                       fg=RED_FG if h_ret.total_seconds() > 0 else '999999',
                       size=8, italic=True, border=b_all())
                    sc(ws.cell(row=base + 6, column=col),
                       value=f"H.sup : {fmt_duree(Pointage(heures_travaillees=h_sup))}" if h_sup.total_seconds() > 0 else '—',
                       bg=GREEN_BG if h_sup.total_seconds() > 0 else bg_day,
                       fg=GREEN_FG if h_sup.total_seconds() > 0 else '999999',
                       size=8, italic=True, border=b_all())
                    sc(ws.cell(row=base + 7, column=col), bg=GREY_LIGHT,
                       border=Border(bottom=Side(style='medium', color=BLUE),
                                     left=sd(), right=sd()))
            
            ws.merge_cells(start_row=base, start_column=COL_TOTAL,
                           end_row=base + 6, end_column=COL_TOTAL)
            
            total_lines = [
                f"Retards :\n{Pointage(heures_travaillees=tot_retard).get_duree_formatee()}" if tot_retard.total_seconds() > 0 else "Retards :\n0h00",
                f"\nH. Travaillées :\n{Pointage(heures_travaillees=tot_trav).get_duree_formatee()}",
                f"\nH. Supp :\n{Pointage(heures_travaillees=tot_sup).get_duree_formatee()}" if tot_sup.total_seconds() > 0 else "\nH. Supp :\n0h00",
            ]
            if tot_gardes > 0:
                total_lines.append(f"\nGardes :\n{tot_gardes}")
            
            sc(ws.cell(row=base, column=COL_TOTAL),
               value="\n".join(total_lines),
               bg=TOTAL_BG, fg=DARK, size=9,
               wrap=True, halign='center', valign='center',
               border=b_outer(BLUE))
            ws.cell(row=base + 7, column=COL_TOTAL).fill = PatternFill('solid', start_color=GREY_MID)
            ws.cell(row=base + 7, column=COL_TOTAL).border = b_bottom()
        
        ws.freeze_panes = f'{get_column_letter(COL_DAYS)}{HEADER_ROW + 1}'
        
        filename = f"resume_pointages_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


# ============================================================
# DEMANDE DE MODIFICATION
# ============================================================

@admin.register(DemandeModification)
class DemandeModificationAdmin(admin.ModelAdmin):
    list_display = ('demandeur', 'type_action', 'cible', 'statut_badge', 'date_creation', 'boutons_action')
    list_filter = ('statut', 'type_action', 'cible')
    search_fields = ('demandeur__username',)
    actions = ['approuver_demandes', 'refuser_demandes']

    readonly_fields = (
        'demandeur', 'type_action', 'cible',
        'donnees_formatees', 'statut_badge',
        'date_creation', 'traitee_par', 'date_traitement',
        'boutons_fiche',
    )

    fieldsets = (
        ('Informations', {
            'fields': ('demandeur', 'type_action', 'cible', 'date_creation')
        }),
        ('Données soumises', {
            'fields': ('donnees_formatees',)
        }),
        ('Statut', {
            'fields': ('statut_badge', 'traitee_par', 'date_traitement')
        }),
        ('Commentaire', {
            'fields': ('commentaire',)
        }),
    )

    def has_add_permission(self, request):
        return False

    def statut_badge(self, obj):
        styles = {
            'en_attente': ('rgba(251,191,36,.12)', '#fbbf24', '⏳ En attente'),
            'approuvee': ('rgba(74,222,128,.12)', '#4ade80', '✅ Approuvée'),
            'refusee': ('rgba(248,113,113,.12)', '#f87171', '❌ Refusée'),
        }
        bg, color, label = styles.get(obj.statut, ('rgba(255,255,255,.07)', '#e8eaf0', obj.statut))
        return mark_safe(
            f'<span style="background:{bg};color:{color};padding:4px 12px;'
            f'border-radius:20px;font-size:12px;font-weight:600;white-space:nowrap;">'
            f'{label}</span>'
        )
    statut_badge.short_description = 'Statut'

    def boutons_action(self, obj):
        if obj.statut != 'en_attente':
            return mark_safe(
                '<span style="color:rgba(232,234,240,.3);font-size:12px;font-style:italic;">Traitée</span>'
            )
        url_approuver = reverse('admin:demande_approuver', args=[obj.pk])
        url_refuser = reverse('admin:demande_refuser', args=[obj.pk])
        return mark_safe(
            f'<div style="display:flex;gap:6px;">'
            f'<a href="{url_approuver}" style="background:rgba(74,222,128,.12);color:#4ade80;'
            f'border:1px solid rgba(74,222,128,.3);padding:4px 12px;border-radius:6px;'
            f'font-size:11px;font-weight:600;text-decoration:none;white-space:nowrap;">✅ Accepter</a>'
            f'<a href="{url_refuser}" style="background:rgba(248,113,113,.12);color:#f87171;'
            f'border:1px solid rgba(248,113,113,.3);padding:4px 12px;border-radius:6px;'
            f'font-size:11px;font-weight:600;text-decoration:none;white-space:nowrap;">❌ Refuser</a>'
            f'</div>'
        )
    boutons_action.short_description = 'Actions'

    def boutons_fiche(self, obj):
        btn_retour = (
            '<a href="../" style="display:inline-flex;align-items:center;gap:6px;'
            'background:rgba(255,255,255,.06);color:#e8eaf0;'
            'border:1px solid rgba(255,255,255,.15);padding:8px 18px;'
            'border-radius:6px;font-size:13px;font-weight:600;text-decoration:none;">'
            '← Retour</a>'
        )
        if obj.statut != 'en_attente':
            return mark_safe(btn_retour)

        return mark_safe(
            '<div style="display:flex;gap:10px;align-items:center;flex-wrap:wrap;">'
            '<button type="submit" name="_accepter" value="1" '
            'style="background:rgba(74,222,128,.15);color:#4ade80;'
            'border:1px solid rgba(74,222,128,.35);padding:8px 20px;'
            'border-radius:6px;font-size:13px;font-weight:600;cursor:pointer;">'
            '✅ Accepter</button>'
            '<button type="submit" name="_refuser" value="1" '
            'style="background:rgba(248,113,113,.15);color:#f87171;'
            'border:1px solid rgba(248,113,113,.35);padding:8px 20px;'
            'border-radius:6px;font-size:13px;font-weight:600;cursor:pointer;">'
            '❌ Refuser</button>'
            + btn_retour +
            '</div>'
        )
    boutons_fiche.short_description = ''

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('<int:pk>/approuver/', self.admin_site.admin_view(self.approuver_view), name='demande_approuver'),
            path('<int:pk>/refuser/', self.admin_site.admin_view(self.refuser_view), name='demande_refuser'),
        ]
        return custom + urls

    def approuver_view(self, request, pk):
        demande = get_object_or_404(DemandeModification, pk=pk)
        if demande.statut == 'en_attente':
            try:
                self._appliquer_demande(demande)
                demande.statut = 'approuvee'
                demande.traitee_par = request.user
                demande.date_traitement = timezone.now()
                demande.save()
                self.message_user(request, f"✅ Demande #{pk} approuvée et appliquée.")
            except Exception as e:
                self.message_user(request, f"❌ Erreur : {e}", level='error')
        return HttpResponseRedirect("../../")

    def refuser_view(self, request, pk):
        demande = get_object_or_404(DemandeModification, pk=pk)
        if demande.statut == 'en_attente':
            demande.statut = 'refusee'
            demande.traitee_par = request.user
            demande.date_traitement = timezone.now()
            demande.save()
            self.message_user(request, f"❌ Demande #{pk} refusée.")
        return HttpResponseRedirect("../../")

    def response_change(self, request, obj):
        if '_accepter' in request.POST and obj.statut == 'en_attente':
            try:
                self._appliquer_demande(obj)
                obj.statut = 'approuvee'
                obj.traitee_par = request.user
                obj.date_traitement = timezone.now()
                obj.save()
                self.message_user(request, f"✅ Demande #{obj.pk} approuvée et appliquée.")
            except Exception as e:
                self.message_user(request, f"❌ Erreur : {e}", level='error')
            return HttpResponseRedirect("../")

        if '_refuser' in request.POST and obj.statut == 'en_attente':
            obj.statut = 'refusee'
            obj.traitee_par = request.user
            obj.date_traitement = timezone.now()
            obj.save()
            self.message_user(request, f"❌ Demande #{obj.pk} refusée.")
            return HttpResponseRedirect("../")

        return super().response_change(request, obj)

    def save_model(self, request, obj, form, change):
        if change:
            original = DemandeModification.objects.get(pk=obj.pk)
            obj.demandeur = original.demandeur
            obj.type_action = original.type_action
            obj.cible = original.cible
            obj.cible_id = original.cible_id
            obj.donnees = original.donnees
            obj.statut = original.statut
            obj.traitee_par = original.traitee_par
            obj.date_traitement = original.date_traitement
        super().save_model(request, obj, form, change)

    _CIBLE_MODELE = {'employe': Employe, 'site': Site, 'poste': Poste}

    def _label_champ(self, modele, champ):
        """Libellé humain d'un champ, tiré du modèle réel (verbose_name) —
        se met à jour tout seul si le modèle change, jamais codé en dur."""
        try:
            return modele._meta.get_field(champ).verbose_name.capitalize()
        except Exception:
            return champ.replace('_', ' ').capitalize()

    def _valeur_affichable(self, champ, valeur):
        """Résout les valeurs de clé étrangère (ex: poste=3) vers un
        libellé lisible (ex: 'Infirmier') plutôt qu'un ID brut."""
        if valeur is None or valeur == '':
            return mark_safe('<span style="color:rgba(255,255,255,.35);font-style:italic;">vide</span>')
        if champ == 'poste':
            poste = Poste.objects.filter(pk=valeur).first()
            return poste.nom if poste else f"#{valeur} (introuvable)"
        return str(valeur)

    def donnees_formatees(self, obj):
        if not obj.donnees and obj.type_action != 'delete':
            return '—'

        modele = self._CIBLE_MODELE.get(obj.cible)
        cible_actuelle = None
        if modele and obj.cible_id and obj.type_action in ('update', 'delete'):
            cible_actuelle = modele.objects.filter(pk=obj.cible_id).first()

        entete_cible = ''
        if obj.type_action != 'create':
            if cible_actuelle is not None:
                entete_cible = format_html(
                    '<div style="padding:8px 14px;background:rgba(79,142,247,.08);'
                    'border-radius:6px;margin-bottom:10px;font-size:12px;color:#e8eaf0;">'
                    'Concerne : <strong>{}</strong></div>',
                    str(cible_actuelle)
                )
            else:
                entete_cible = mark_safe(
                    '<div style="padding:8px 14px;background:rgba(248,113,113,.1);'
                    'border-radius:6px;margin-bottom:10px;font-size:12px;color:#f87171;">'
                    "⚠️ L'élément visé n'existe plus (déjà supprimé ou modifié depuis)</div>"
                )

        if obj.type_action == 'delete':
            if not cible_actuelle:
                return mark_safe(entete_cible or '—')
            lignes = []
            for champ in modele._meta.fields:
                if champ.name in ('id',):
                    continue
                lignes.append(format_html(
                    '<tr><td style="padding:8px 14px;color:rgba(255,255,255,.45);font-size:11px;'
                    'text-transform:uppercase;letter-spacing:.08em;white-space:nowrap;'
                    'border-bottom:1px solid rgba(255,255,255,.06)">{}</td>'
                    '<td style="padding:8px 14px;font-weight:500;color:#e8eaf0;'
                    'border-bottom:1px solid rgba(255,255,255,.06)">{}</td></tr>',
                    self._label_champ(modele, champ.name),
                    str(getattr(cible_actuelle, champ.name))
                ))
            return format_html(
                '{}<table style="border-collapse:collapse;width:100%;background:#1c2236;'
                'border-radius:8px;overflow:hidden;border:1px solid rgba(255,255,255,.07)">'
                '<tr><td colspan="2" style="padding:8px 14px;font-size:11px;color:#f87171;'
                'text-transform:uppercase;letter-spacing:.08em;">🗑️ Élément qui sera supprimé</td></tr>{}</table>',
                mark_safe(entete_cible), mark_safe(''.join(lignes))
            )

        lignes = []
        for champ, nouvelle_valeur in obj.donnees.items():
            label = self._label_champ(modele, champ) if modele else champ.replace('_', ' ').capitalize()
            nouvelle_affichee = self._valeur_affichable(champ, nouvelle_valeur)

            if obj.type_action == 'update' and cible_actuelle is not None:
                ancienne_brute = getattr(cible_actuelle, f"{champ}_id", None)
                if ancienne_brute is None and not hasattr(cible_actuelle, f"{champ}_id"):
                    ancienne_brute = getattr(cible_actuelle, champ, None)
                ancienne_affichee = self._valeur_affichable(champ, ancienne_brute)
                a_change = str(ancienne_brute) != str(nouvelle_valeur)

                lignes.append(format_html(
                    '<tr style="{}">'
                    '<td style="padding:8px 14px;color:rgba(255,255,255,.45);font-size:11px;'
                    'text-transform:uppercase;letter-spacing:.08em;white-space:nowrap;'
                    'border-bottom:1px solid rgba(255,255,255,.06)">{}</td>'
                    '<td style="padding:8px 14px;color:{};border-bottom:1px solid rgba(255,255,255,.06)">{}</td>'
                    '<td style="padding:8px 14px;text-align:center;color:rgba(255,255,255,.3);'
                    'border-bottom:1px solid rgba(255,255,255,.06)">→</td>'
                    '<td style="padding:8px 14px;font-weight:600;color:{};'
                    'border-bottom:1px solid rgba(255,255,255,.06)">{}</td></tr>',
                    'background:rgba(74,222,128,.05);' if a_change else '',
                    label,
                    'rgba(255,255,255,.35);text-decoration:line-through;' if a_change else '#e8eaf0;',
                    ancienne_affichee,
                    '#4ade80;' if a_change else '#e8eaf0;',
                    nouvelle_affichee,
                ))
            else:
                # Création (pas de "avant" — l'élément n'existe pas encore)
                lignes.append(format_html(
                    '<tr>'
                    '<td style="padding:8px 14px;color:rgba(255,255,255,.45);font-size:11px;'
                    'text-transform:uppercase;letter-spacing:.08em;white-space:nowrap;'
                    'border-bottom:1px solid rgba(255,255,255,.06)">{}</td>'
                    '<td colspan="3" style="padding:8px 14px;font-weight:500;color:#e8eaf0;'
                    'border-bottom:1px solid rgba(255,255,255,.06)">{}</td></tr>',
                    label, nouvelle_affichee
                ))

        entete_colonnes = '' if obj.type_action == 'create' else (
            '<tr><td></td><td style="padding:4px 14px;font-size:10px;color:rgba(255,255,255,.35);'
            'text-transform:uppercase;">Avant</td><td></td>'
            '<td style="padding:4px 14px;font-size:10px;color:rgba(255,255,255,.35);'
            'text-transform:uppercase;">Après</td></tr>'
        )
        return format_html(
            '{}<table style="border-collapse:collapse;width:100%;background:#1c2236;'
            'border-radius:8px;overflow:hidden;border:1px solid rgba(255,255,255,.07)">{}{}</table>',
            mark_safe(entete_cible), mark_safe(entete_colonnes), mark_safe(''.join(lignes))
        )
    donnees_formatees.short_description = "Détails de la demande"

    @admin.action(description="✅ Approuver les demandes sélectionnées")
    def approuver_demandes(self, request, queryset):
        for demande in queryset.filter(statut='en_attente'):
            try:
                self._appliquer_demande(demande)
                demande.statut = 'approuvee'
                demande.traitee_par = request.user
                demande.date_traitement = timezone.now()
                demande.save()
            except Exception as e:
                self.message_user(request, f"❌ Erreur demande #{demande.pk} : {e}", level='error')
        self.message_user(request, "✅ Demandes approuvées et appliquées.")

    @admin.action(description="❌ Refuser les demandes sélectionnées")
    def refuser_demandes(self, request, queryset):
        for demande in queryset.filter(statut='en_attente'):
            demande.statut = 'refusee'
            demande.traitee_par = request.user
            demande.date_traitement = timezone.now()
            demande.save()
        self.message_user(request, "❌ Demandes refusées.")

    def _appliquer_demande(self, demande):
        d = demande.donnees

        if demande.cible == 'employe':
            if demande.type_action == 'create':
                Employe.objects.create(
                    nom=d['nom'], prenom=d['prenom'],
                    matricule=d['matricule'],
                    poste_id=d.get('poste'),
                    actif=d.get('actif', True)
                )
            elif demande.type_action == 'update':
                Employe.objects.filter(pk=demande.cible_id).update(
                    nom=d['nom'], prenom=d['prenom'],
                    matricule=d['matricule'],
                    poste_id=d.get('poste'),
                    actif=d.get('actif', True)
                )
            elif demande.type_action == 'delete':
                try:
                    employe = Employe.objects.get(pk=demande.cible_id)
                    employe.delete()
                except Employe.DoesNotExist:
                    pass

        elif demande.cible == 'site':
            if demande.type_action == 'create':
                Site.objects.create(
                    nom=d['nom'], adresse=d.get('adresse', ''),
                    heure_ouverture_matin=d['heure_ouverture_matin'],
                    heure_fermeture_matin=d['heure_fermeture_matin'],
                    heure_ouverture_apres_midi=d['heure_ouverture_apres_midi'],
                    heure_fermeture_apres_midi=d['heure_fermeture_apres_midi'],
                )
            elif demande.type_action == 'update':
                Site.objects.filter(pk=demande.cible_id).update(
                    nom=d['nom'], adresse=d.get('adresse', ''),
                    heure_ouverture_matin=d['heure_ouverture_matin'],
                    heure_fermeture_matin=d['heure_fermeture_matin'],
                    heure_ouverture_apres_midi=d['heure_ouverture_apres_midi'],
                    heure_fermeture_apres_midi=d['heure_fermeture_apres_midi'],
                )
            elif demande.type_action == 'delete':
                Site.objects.filter(pk=demande.cible_id).delete()

        elif demande.cible == 'poste':
            if demande.type_action == 'create':
                Poste.objects.create(
                    nom=d['nom'],
                    description=d.get('description', ''),
                    couleur=d.get('couleur', '#4361ee')
                )
            elif demande.type_action == 'update':
                Poste.objects.filter(pk=demande.cible_id).update(
                    nom=d['nom'],
                    description=d.get('description', ''),
                    couleur=d.get('couleur', '#4361ee')
                )
            elif demande.type_action == 'delete':
                Poste.objects.filter(pk=demande.cible_id).delete()


# ============================================================
# SCAN
# ============================================================

@admin.register(Scan)
class ScanAdmin(admin.ModelAdmin):
    list_display = ('employe', 'site', 'timestamp_local', 'type_scan_display')
    list_filter = ('type_scan', 'site', 'timestamp')
    search_fields = ('employe__nom', 'employe__prenom', 'employe__matricule')
    readonly_fields = ('timestamp',)
    date_hierarchy = 'timestamp'

    def timestamp_local(self, obj):
        return obj.get_timestamp_local().strftime('%d/%m/%Y %H:%M:%S')
    timestamp_local.short_description = 'Heure locale'

    def type_scan_display(self, obj):
        colors = {
            'entree_matin': '#4f8ef7',
            'sortie_matin': '#4ade80',
            'entree_apres_midi': '#fbbf24',
            'sortie_apres_midi': '#22d3ee',
            'debut_garde': '#a78bfa',
            'fin_garde': '#94a3b8',
        }
        color = colors.get(obj.type_scan, '#94a3b8')
        return mark_safe(
            f'<span style="background:rgba(255,255,255,.07);color:{color};'
            f'padding:3px 10px;border-radius:20px;font-size:11px;font-weight:600;'
            f'letter-spacing:.05em;border:1px solid {color}40">'
            f'{obj.get_type_scan_display()}</span>'
        )
    type_scan_display.short_description = 'Type de scan'

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('employe', 'site')


# ============================================================
# ANOMALIES DE POINTAGE
# ============================================================

class AnomalieTraitementInline(admin.StackedInline):
    model = AnomalieTraitement
    can_delete = False
    extra = 0
    max_num = 1
    fields = ('administrateur', 'date_traitement', 'commentaire', 'pointage_concerne', 'corrections')
    readonly_fields = ('administrateur', 'date_traitement')


@admin.register(AnomaliePointage)
class AnomaliePointageAdmin(admin.ModelAdmin):
    list_display = ('type_display', 'employe_ou_matricule', 'gravite_badge', 'statut_badge', 'created_at')
    list_filter = ('statut', 'type')
    search_fields = ('employe__nom', 'employe__prenom', 'employe__matricule', 'matricule_scanne', 'message')
    date_hierarchy = 'created_at'
    inlines = [AnomalieTraitementInline]
    actions = ['marquer_traitees', 'marquer_cloturees']
    ordering = ('-created_at',)

    readonly_fields = (
        'type', 'employe', 'matricule_scanne', 'site',
        'message', 'contexte_formate', 'gravite_badge', 'statut_badge',
        'cloturee_par', 'date_cloture', 'created_at',
    )

    fieldsets = (
        ("Anomalie", {
            'fields': ('type', 'gravite_badge', 'employe', 'matricule_scanne', 'site', 'created_at')
        }),
        ("Détails", {'fields': ('message', 'contexte_formate')}),
        ("Statut", {'fields': ('statut_badge', 'cloturee_par', 'date_cloture')}),
    )

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def type_display(self, obj):
        return obj.get_type_display()
    type_display.short_description = 'Type'

    def employe_ou_matricule(self, obj):
        if obj.employe:
            return obj.employe.get_nom_complet()
        return obj.matricule_scanne or '—'
    employe_ou_matricule.short_description = 'Employé'

    def gravite_badge(self, obj):
        styles = {
            'info': ('rgba(96,165,250,.12)', '#60a5fa', 'ℹ️ Info'),
            'warning': ('rgba(251,191,36,.12)', '#fbbf24', '⚠️ Avertissement'),
            'critique': ('rgba(248,113,113,.12)', '#f87171', '🚨 Critique'),
        }
        bg, color, label = styles.get(obj.gravite, ('rgba(255,255,255,.07)', '#e8eaf0', obj.gravite))
        return mark_safe(
            f'<span style="background:{bg};color:{color};padding:4px 12px;'
            f'border-radius:20px;font-size:12px;font-weight:600;white-space:nowrap;">'
            f'{label}</span>'
        )
    gravite_badge.short_description = 'Gravité'

    def statut_badge(self, obj):
        styles = {
            AnomaliePointage.STATUT_OUVERTE: ('rgba(248,113,113,.12)', '#f87171', '🔴 Ouverte'),
            AnomaliePointage.STATUT_TRAITEE: ('rgba(251,191,36,.12)', '#fbbf24', '🟡 Traitée'),
            AnomaliePointage.STATUT_CLOTUREE: ('rgba(74,222,128,.12)', '#4ade80', '✅ Clôturée'),
        }
        bg, color, label = styles.get(obj.statut, ('rgba(255,255,255,.07)', '#e8eaf0', obj.statut))
        return mark_safe(
            f'<span style="background:{bg};color:{color};padding:4px 12px;'
            f'border-radius:20px;font-size:12px;font-weight:600;white-space:nowrap;">'
            f'{label}</span>'
        )
    statut_badge.short_description = 'Statut'

    def contexte_formate(self, obj):
        if not obj.contexte:
            return '—'
        return format_html(
            '<pre style="background:#1c2236;border:1px solid rgba(255,255,255,.07);'
            'border-radius:8px;padding:12px 14px;font-size:12px;color:#e8eaf0;'
            'white-space:pre-wrap;">{}</pre>',
            json.dumps(obj.contexte, indent=2, ensure_ascii=False, default=str)
        )
    contexte_formate.short_description = "Contexte"

    @admin.action(description="✅ Marquer comme traitées")
    def marquer_traitees(self, request, queryset):
        count = 0
        for anomalie in queryset.exclude(statut=AnomaliePointage.STATUT_CLOTUREE):
            try:
                marquer_traitee(
                    anomalie, request.user,
                    type_action=AnomalieTraitement.ACTION_CORRECTION,
                    commentaire="Marquée traitée depuis l'administration.",
                )
                count += 1
            except ValueError as e:
                self.message_user(request, f"❌ Anomalie #{anomalie.pk} : {e}", level=messages.ERROR)
        self.message_user(request, f"✅ {count} anomalie(s) marquée(s) comme traitée(s).")

    @admin.action(description="🔒 Clôturer")
    def marquer_cloturees(self, request, queryset):
        count = 0
        for anomalie in queryset:
            try:
                marquer_cloturee(anomalie, request.user)
                count += 1
            except ValueError as e:
                self.message_user(request, f"❌ Anomalie #{anomalie.pk} : {e}", level=messages.ERROR)
        self.message_user(request, f"🔒 {count} anomalie(s) clôturée(s).")

    # ------------------------------------------------------------
    # Correction réelle du pointage depuis l'admin d'une anomalie.
    # Réutilise PointageForm + marquer_traitee() (déjà existants) —
    # écrit directement le Pointage, ne passe jamais par process_scan()
    # ni par la state machine (c'est une correction manuelle RH, pas
    # un scan).
    # ------------------------------------------------------------
    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                '<int:anomalie_id>/corriger-pointage/',
                self.admin_site.admin_view(self.corriger_pointage_view),
                name='anomalie_corriger_pointage',
            ),
        ]
        return custom + urls

    def corriger_pointage_view(self, request, anomalie_id):
        anomalie = get_object_or_404(AnomaliePointage, pk=anomalie_id)

        if anomalie.statut == AnomaliePointage.STATUT_CLOTUREE:
            self.message_user(request, "❌ Cette anomalie est déjà clôturée.", level=messages.ERROR)
            return redirect(f'/admin/pointage/anomaliepointage/{anomalie.pk}/change/')

        if request.method == 'POST':
            commentaire = request.POST.get('commentaire', '').strip()
            employe_id = request.POST.get('employe')
            date_pointage = request.POST.get('date_pointage')
            periode = request.POST.get('periode')

            pointage_existant = Pointage.objects.filter(
                employe_id=employe_id, date_pointage=date_pointage, periode=periode
            ).first()

            # Capturer les anciennes valeurs AVANT is_valid() : ModelForm
            # mute l'instance liée pendant _post_clean().
            anciennes_valeurs = {}
            if pointage_existant:
                for champ in ['site', 'type_journee', 'heure_arrivee', 'heure_depart', 'statut', 'notes']:
                    anciennes_valeurs[champ] = getattr(pointage_existant, champ)

            form = PointageForm(request.POST, instance=pointage_existant)
            if not form.is_valid():
                return render(request, 'admin/pointage/anomalie/corriger_pointage.html', {
                    'anomalie': anomalie, 'form': form, 'opts': self.model._meta,
                })

            cd = form.cleaned_data
            corrections = []
            created = pointage_existant is None
            if not created:
                for champ, ancienne in anciennes_valeurs.items():
                    nouvelle = cd[champ]
                    if ancienne != nouvelle:
                        corrections.append({
                            'champ': champ,
                            'ancienne_valeur': str(ancienne) if ancienne is not None else None,
                            'nouvelle_valeur': str(nouvelle) if nouvelle is not None else None,
                        })

            pointage = form.save()

            try:
                marquer_traitee(
                    anomalie, request.user, type_action=AnomalieTraitement.ACTION_CORRECTION,
                    commentaire=commentaire, corrections=corrections, pointage_concerne=pointage,
                )
                self.message_user(request, f"✅ Pointage {'créé' if created else 'corrigé'}, anomalie #{anomalie.pk} traitée.")
            except (ValueError, PermissionError) as e:
                self.message_user(request, f"❌ {e}", level=messages.ERROR)

            return redirect(f'/admin/pointage/anomaliepointage/{anomalie.pk}/change/')

        # GET : préremplir avec le pointage existant s'il y en a un
        initial = {}
        pointage_existant = None
        if anomalie.employe and anomalie.date_pointage:
            pointage_existant = Pointage.objects.filter(
                employe=anomalie.employe, date_pointage=anomalie.date_pointage,
            ).order_by('periode').first()
        if pointage_existant:
            initial = {
                'employe': pointage_existant.employe_id, 'site': pointage_existant.site_id,
                'date_pointage': pointage_existant.date_pointage, 'periode': pointage_existant.periode,
                'type_journee': pointage_existant.type_journee,
                'heure_arrivee': pointage_existant.heure_arrivee, 'heure_depart': pointage_existant.heure_depart,
                'statut': pointage_existant.statut, 'notes': pointage_existant.notes,
            }
        elif anomalie.employe:
            initial = {'employe': anomalie.employe_id, 'site': anomalie.site_id, 'date_pointage': anomalie.date_pointage}

        form = PointageForm(initial=initial)
        return render(request, 'admin/pointage/anomalie/corriger_pointage.html', {
            'anomalie': anomalie, 'form': form, 'opts': self.model._meta,
        })


# ============================================================
# CONFIGURATION DU SITE ADMIN
# ============================================================

admin.site.site_header = "Pointage QR — Administration"
admin.site.site_title = "Pointage QR"
admin.site.index_title = "Tableau de bord d'administration"