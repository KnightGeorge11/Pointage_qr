# pointage/tests/test_export_excel.py
#
# Non-régression : matin.get_retard_minutes (sans parenthèses) est une
# méthode, pas une @property -- l'appeler sans () renvoyait un objet
# "bound method" stringifié dans la cellule Excel au lieu du nombre de
# minutes de retard réel.

import io
from datetime import time as dtime, date

from django.test import TestCase, Client
from django.urls import reverse
from openpyxl import load_workbook

from pointage.models import CustomUser, Employe, Site, Pointage


class TestExportExcelRetard(TestCase):
    def setUp(self):
        self.admin = CustomUser.objects.create_user(
            username="admin_export", password="pass1234", role="admin", is_staff=True,
        )
        self.employe = Employe.objects.create(nom="Rakoto", prenom="Jean", matricule="E020", actif=True)
        self.site = Site.objects.create(
            nom="Site", adresse="a",
            heure_ouverture_matin=dtime(8, 0), heure_fermeture_matin=dtime(12, 0),
            heure_ouverture_apres_midi=dtime(13, 30), heure_fermeture_apres_midi=dtime(17, 30),
        )
        # Entrée à 8h20 sur un site ouvrant à 8h00 -> 20 minutes de retard
        self.pointage = Pointage.objects.create(
            employe=self.employe, site=self.site, date_pointage=date.today(),
            periode='matin', type_journee='normal',
            heure_arrivee=dtime(8, 20), heure_depart=dtime(12, 0),
        )
        self.client = Client()
        self.client.force_login(self.admin)

    def test_export_ne_contient_jamais_un_repr_de_methode(self):
        response = self.client.get(reverse('export_resume_excel'))
        assert response.status_code == 200

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        contenu_complet = "\n".join(
            str(cell.value) for row in ws.iter_rows() for cell in row if cell.value
        )
        assert 'bound method' not in contenu_complet
        assert 'get_retard_minutes' not in contenu_complet

    def test_export_affiche_bien_un_retard_numerique(self):
        response = self.client.get(reverse('export_resume_excel'))
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        cellules_retard = [
            cell.value for row in ws.iter_rows() for cell in row
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('Retard :')
        ]
        assert len(cellules_retard) >= 1
        # Le retard réel (matin: 8h20 vs ouverture 8h00) doit apparaître
        # comme un nombre suivi de "min", pas comme un repr Python.
        assert any('20min' in c for c in cellules_retard), cellules_retard
